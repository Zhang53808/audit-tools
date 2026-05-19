#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
固定资产折旧测算脚本 —— 通用版

每换一个客户，第一次运行时会自动识别列名。
识别不对的话，把表头发给我，我告诉你配置文件怎么写。
确认后保存配置，下次同客户直接调用。
"""

import openpyxl, json, os
from datetime import datetime
from collections import defaultdict

# 常用关键词映射（自动匹配用）
AUTO_MAP = {
    "card": ["卡片编码", "卡片编号", "资产编码", "资产编号", "固定资产编码"],
    "name": ["资产名称", "固定资产名称", "设备名称"],
    "entry_date": ["入账日期", "入账日", "启用日期", "开始使用日期", "购入日期"],
    "life": ["使用寿命", "使用年限", "折旧年限", "折旧期限", "预计使用月数", "预计使用年限"],
    "used_life": ["已使用寿命", "已使用月数", "已提月数", "已使用年限", "累计已提月数"],
    "residual": ["残值", "预计净残值", "净残值", "残值额"],
    "residual_rate": ["残值率", "净残值率", "预计净残值率"],
    "orig_value": ["期初原值", "原值", "资产原值", "固定资产原值", "原值(期初)", "原值期初"],
    "monthly_depr": ["本期折旧额", "月折旧额", "本期折旧", "当月折旧"],
    "monthly_rate": ["月折旧率"],
    "orig_increase": ["原值增加", "原值调增", "本期新增原值"],  # TODO: 预留，本期增加资产的折旧起算逻辑
    "orig_decrease": ["原值减少", "原值调减", "本期减少原值"],
}

print("请把客户的固定资产明细表拖进终端，然后回车：")
path = input().strip().strip("'\"")

wb = openpyxl.load_workbook(path)
ws = wb.active

# 配置路径
config_dir = os.path.join(os.path.dirname(__file__) or ".", "客户配置")
os.makedirs(config_dir, exist_ok=True)
customer_name = os.path.splitext(os.path.basename(path))[0]
config_path = os.path.join(config_dir, f"{customer_name}.json")

# 检查是否有已保存的配置
loaded_config = False
if os.path.exists(config_path):
    yn = input(f"发现已保存的配置（{customer_name}.json），直接使用？(y/n)：").strip().lower()
    if yn == "y":
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
        data_start = config.get("data_start", 2)
        is_grouped = config.get("grouped", False)
        audit_year = config.get("audit_year", 2025)
        config["method"] = config.get("method", "straight")
        config["has_disposal"] = config.get("has_disposal", False)
        print(f"已加载配置：数据第{data_start}行开始，分组={is_grouped}，年度={audit_year}")
        loaded_config = True

if not loaded_config:
    # 读取第一行表头
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        vals = [str(v) if v else "" for v in row]
        non_empty = [v for v in vals if v]
        # 找到有足够多非空列且包含资产相关关键词的行
        text = " ".join(non_empty)
        keywords = ["资产", "卡片", "原值", "折旧", "残值", "入账"]
        if sum(1 for k in keywords if k in text) >= 3:
            header_row = vals
            break

    if header_row is None:
        print("找不到表头行，请确认文件格式")
        exit(1)

    print(f"\n读取到 {len(header_row)} 列，表头如下：")
    for j, h in enumerate(header_row):
        if h:
            print(f"  列{j}: {h}")

    # 自动匹配列位置
    config = {}
    unmatched = []
    for field, keywords in AUTO_MAP.items():
        found = False
        for j, h in enumerate(header_row):
            if any(k in h for k in keywords):
                config[field] = j
                found = True
                break
        if not found:
            unmatched.append(field)

    print(f"\n--- 自动匹配结果 ---")
    if config:
        for field, col in sorted(config.items(), key=lambda x: x[1]):
            print(f"  ✅ {field} → 列{col} ({header_row[col]})")
    if unmatched:
        print(f"\n  ❌ 以下字段未匹配到：{', '.join(unmatched)}")
        print("     请你把上面列出来的表头发给我，我告诉你列号填什么")

    # 手动补充
    if unmatched:
        print("\n手动补充未匹配的列号：")
        for field in unmatched:
            try:
                val = input(f"  {field} 在第几列？(直接回车跳过)：").strip()
                if val:
                    config[field] = int(val)
            except:
                pass

    # 确认数据起始行
    data_start = input(f"\n数据从第几行开始？(默认第2行)：").strip()
    data_start = int(data_start) if data_start.isdigit() else 2

    # 是否按月分组
    is_grouped = input("\n同一个资产会出现在多行吗？（按月列示的选y，一行一条的选n）[y/n]：").strip().lower() == "y"

    # 审计年度
    audit_year = input("\n审计年度？(默认2025)：").strip()
    audit_year = int(audit_year) if audit_year.isdigit() else 2025

    print(f"\n--- 配置确认 ---")
    config["data_start"] = data_start
    config["grouped"] = is_grouped
    config["audit_year"] = audit_year

    save = input(f"\n保存配置以便下次直接使用？(y/n)：").strip().lower() == "y"
    if save:
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        print(f"已保存：{config_path}")

    # 折旧方法
    print("\n折旧方法？(按回车默认直线法)")
    print("  1 - 直线法（原值-残值）/寿命")
    print("  2 - 按系统折旧率（原值×月折旧率）")
    depr_method = input("请选择 1 或 2 (默认1)：").strip()
    if depr_method == "2":
        config["method"] = "rate"
    else:
        config["method"] = "straight"

    # 是否有处置
    has_disp = input("\n本年度是否有资产处置/报废？(y/n，默认n)：").strip().lower() == "y"
    if has_disp and config.get("orig_decrease"):
        config["has_disposal"] = True
    else:
        config["has_disposal"] = False

# ========== 开始计算 ==========
print(f"\n正在测算 {audit_year} 年折旧...\n")

rows_data = list(ws.iter_rows(min_row=data_start, values_only=True))

# 年份校验：检查入账日期是否跨年
if config.get("entry_date") is not None:
    years_found = set()
    for row in rows_data:
        ed = row[config["entry_date"]]
        if isinstance(ed, datetime):
            years_found.add(ed.year)
        elif isinstance(ed, str):
            for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"]:
                try:
                    years_found.add(datetime.strptime(ed.strip(), fmt).year)
                    break
                except:
                    pass
    if len(years_found) > 1:
        print(f"⚠️  警告：入账日期包含多个年份 {sorted(years_found)}")
        print(f"   数据可能未经筛选，min(已提月数)可能不是年初值")
        yn = input("   继续？(y/n)：").strip().lower()
        if yn != "y":
            exit(0)

if is_grouped:
    # 按月分组模式（同一资产多行）
    asset_groups = defaultdict(lambda: {
        "name": "",
        "entry_date": None,
        "life": 0,
        "used_life_min": None,
        "residual": 0,
        "orig_value": 0,
        "total_monthly": 0,
        "monthly_rate": 0,
        "rows": 0,
        "disposed": False,
        "disposal_month": 0
    })

    for row in rows_data:
        card = str(row[config.get("card", 0)] or "").strip()
        name = str(row[config.get("name", 3)] or "").strip()
        if not card and not name:
            continue
        g = asset_groups[card]
        g["name"] = name
        g["entry_date"] = row[config.get("entry_date", 12)]
        try:
            g["life"] = float(row[config["life"]]) if row[config["life"]] else 0
            used_val = float(row[config["used_life"]]) if config.get("used_life") is not None and row[config["used_life"]] else 0
            g["orig_value"] = float(row[config["orig_value"]]) if row[config["orig_value"]] else 0
            g["residual"] = float(row[config["residual"]]) if config.get("residual") is not None and row[config["residual"]] else 0
            # 如果残值额为0但有残值率，用残值率算
            if g["residual"] == 0 and config.get("residual_rate") is not None:
                rate = float(row[config["residual_rate"]]) if row[config["residual_rate"]] else 0
                if rate > 0:
                    g["residual"] = round(g["orig_value"] * rate / 100, 2)
            # 读取月折旧率（按折旧率算法用）
            if config.get("monthly_rate") is not None:
                g["monthly_rate"] = float(row[config["monthly_rate"]]) if row[config["monthly_rate"]] else 0
            monthly = float(row[config["monthly_depr"]]) if config.get("monthly_depr") is not None and row[config["monthly_depr"]] else 0
        except:
            continue
        if g["used_life_min"] is None or used_val < g["used_life_min"]:
            g["used_life_min"] = used_val
        g["total_monthly"] += monthly
        g["rows"] += 1
        
        # 检测原值减少（处置标记）
        if config.get("orig_decrease") is not None:
            try:
                dec = float(row[config["orig_decrease"]]) if row[config["orig_decrease"]] else 0
                if dec > 0:
                    g["disposed"] = True
                    g["disposal_month"] = g["rows"]  # 用当前行数作为处置月份
            except:
                pass

    items = [(card, g) for card, g in sorted(asset_groups.items())]

else:
    # 一行一条模式
    items = []
    for i, row in enumerate(rows_data):
        card = str(row[config.get("card", 0)] or "").strip()
        name = str(row[config.get("name", 3)] or "").strip()
        if not card and not name:
            continue
        items.append((card, {
            "name": name,
            "entry_date": row[config.get("entry_date", 12)],
            "life": float(row[config["life"]]) if row[config["life"]] else 0,
            "used_life_min": float(row[config["used_life"]]) if config.get("used_life") is not None and row[config["used_life"]] else 0,
            "orig_value": float(row[config["orig_value"]]) if row[config["orig_value"]] else 0,
            "residual": float(row[config["residual"]]) if config.get("residual") is not None and row[config["residual"]] else 0,
            "residual_rate": float(row[config["residual_rate"]]) if config.get("residual_rate") is not None and row[config["residual_rate"]] else 0,
            "monthly_rate": float(row[config["monthly_rate"]]) if config.get("monthly_rate") is not None and row[config["monthly_rate"]] else 0,
            "total_monthly": float(row[config["monthly_depr"]]) if config.get("monthly_depr") is not None and row[config["monthly_depr"]] else 0,
            "rows": 1,
            "disposed": False
        }))
        
        # 一行模式检测处置
        if config.get("orig_decrease") is not None:
            try:
                dec = float(row[config["orig_decrease"]]) if row[config["orig_decrease"]] else 0
                if dec > 0:
                    items[-1][1]["disposed"] = True
            except:
                pass

# 测算
results = []
match = 0
mismatch = 0
zero_value = 0
no_date = 0
invalid_life = 0

for card, g in items:
    name = g["name"]
    life = g["life"]
    used_life_min = g["used_life_min"]
    residual = g["residual"]
    orig_value = g["orig_value"]
    
    if residual == 0 and config.get("residual_rate") is not None:
        rate = g.get("residual_rate", 0)
        if rate > 0:
            residual = round(orig_value * rate / 100, 2)
    cust_total = round(g["total_monthly"], 2)

    if orig_value == 0:
        zero_value += 1
        continue
    if life <= 0:
        invalid_life += 1
        continue

    entry_date = g["entry_date"]
    acq_year = audit_year
    acq_month = 1

    if isinstance(entry_date, datetime):
        acq_year = entry_date.year
        acq_month = entry_date.month
        date_parsed = True
    elif isinstance(entry_date, str):
        date_parsed = False
        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%Y-%m", "%Y/%m", "%Y年%m月"]:
            try:
                dt = datetime.strptime(entry_date.strip(), fmt)
                acq_year = dt.year
                acq_month = dt.month
                date_parsed = True
                break
            except:
                pass
        if not date_parsed:
            no_date += 1
            continue
    else:
        no_date += 1
        continue

    # 统一日期格式
    if isinstance(entry_date, datetime):
        entry_str = entry_date.strftime("%Y-%m-%d")
    elif isinstance(entry_date, str):
        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%Y-%m", "%Y/%m", "%Y年%m月"]:
            try:
                dt = datetime.strptime(entry_date.strip(), fmt)
                entry_str = dt.strftime("%Y-%m-%d")
                break
            except:
                pass
        else:
            entry_str = entry_date
    else:
        entry_str = str(entry_date) if entry_date else ""

    # 计算月折旧额
    if config.get("method") == "rate" and config.get("monthly_rate"):
        # 从月折旧率列读取
        monthly_rate = float(g.get("monthly_rate", 0))
        monthly_depr = round(orig_value * monthly_rate / 100, 2)
    else:
        monthly_depr = round((orig_value - residual) / life, 2)  # 直线法

    if used_life_min is not None and used_life_min > 0:
        remaining_at_start = life - (used_life_min - 1)
    else:
        remaining_at_start = life

    if remaining_at_start <= 0:
        calc_total = 0
        months = 0
        tag = "已提足"
    else:
        tag = ""
        if acq_year < audit_year:
            months = 12
        elif acq_year == audit_year:
            months = 12 - acq_month
            if months < 0:
                months = 0
        else:
            continue

        if months > remaining_at_start:
            months = int(remaining_at_start)
        # 处置月份截断
        if config.get("has_disposal") and g.get("disposed"):
            # 分组模式有具体月份
            if g.get("disposal_month", 0) > 0 and g["disposal_month"] < months:
                months = g["disposal_month"]
            else:
                # 一行模式无法确定月份，标记给用户
                if tag:
                    tag += ",有处置"
                else:
                    tag = "有处置"
        if months <= 0:
            continue
        # 用round后的月折旧额算年折旧
        calc_total = round(monthly_depr * months, 2)
        if not g.get("disposed"):
            tag = ""

    diff = round(calc_total - cust_total, 2)
    if tag:
        status = tag + (",一致" if abs(diff) < 0.01 else ",有差异")
    else:
        status = "一致" if abs(diff) < 0.01 else "有差异"
    if "一致" in status:
        match += 1
    else:
        mismatch += 1

    cust_months = g["rows"]
    month_status = "一致" if cust_months == months else f"应提{months}月,客户{cust_months}月"

    results.append([card, name, entry_str, orig_value, residual, int(life),
                   int(used_life_min) if used_life_min else "", cust_months, months,
                   monthly_depr, calc_total, cust_total, diff, status, month_status])

# 输出
out_path = path.replace(".xlsx", "_折旧年审结果.xlsx")
out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = "折旧测算"

headers_list = ["卡片编码", "资产名称", "入账日期", "期初原值", "残值", "使用寿命(月)",
                "年初已提月数", "客户实际月数", "应提月数", "月折旧额",
                "测算年折旧额", "客户年折旧额(汇总)", "差异", "结论", "月数核对"]

for j, h in enumerate(headers_list):
    out_ws.cell(row=1, column=j+1, value=h)

# 冻结首行 + 自动筛选
out_ws.freeze_panes = "A2"
# 自动筛选
from openpyxl.utils import get_column_letter
last_col = get_column_letter(len(headers_list))
out_ws.auto_filter.ref = f"A1:{last_col}{len(results)+1}"

red_font = openpyxl.styles.Font(color="FF0000")
for i, row in enumerate(results, 2):
    for j, v in enumerate(row):
        cell = out_ws.cell(row=i, column=j+1, value=v)
        if "差异" in str(row[-2]):
            cell.font = red_font

# 差异汇总sheet
diff_ws = out_wb.create_sheet(title="差异汇总")
diff_ws.cell(row=1, column=1, value="卡片编码")
diff_ws.cell(row=1, column=2, value="资产名称")
diff_ws.cell(row=1, column=3, value="测算年折旧额")
diff_ws.cell(row=1, column=4, value="客户年折旧额")
diff_ws.cell(row=1, column=5, value="差异")
diff_ws.cell(row=1, column=6, value="备注")

diff_row = 2
total_diff = 0
for row in results:
    if "有差异" in str(row[-2]):
        diff_ws.cell(row=diff_row, column=1, value=row[0])  # 卡片
        diff_ws.cell(row=diff_row, column=2, value=row[1])  # 名称
        diff_ws.cell(row=diff_row, column=3, value=row[10])  # 测算年折旧
        diff_ws.cell(row=diff_row, column=4, value=row[11])  # 客户年折旧
        diff_ws.cell(row=diff_row, column=5, value=row[12])  # 差异
        diff_ws.cell(row=diff_row, column=6, value=row[13])  # 结论
        total_diff += abs(row[12]) if row[12] else 0
        diff_row += 1

diff_ws.cell(row=diff_row + 1, column=1, value="差异合计")
diff_ws.cell(row=diff_row + 1, column=5, value=round(total_diff, 2))
diff_ws.cell(row=diff_row + 1, column=1).font = openpyxl.styles.Font(bold=True)
diff_ws.cell(row=diff_row + 1, column=5).font = openpyxl.styles.Font(bold=True)
diff_ws.freeze_panes = "A2"

out_wb.save(out_path)

print(f"\n完成！结果：{out_path}")
print(f"共处理 {len(results)} 条资产")
print(f"  一致：{match}  有差异：{mismatch}  原值0：{zero_value}  无日期：{no_date}  寿命异常：{invalid_life}")
