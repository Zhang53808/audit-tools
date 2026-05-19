#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成一份模拟的烂格式抽凭清单，用来练习Power Query清洗
运行后会在桌面生成一个Excel文件
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import random
from datetime import datetime, timedelta

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "抽凭清单"

# ---- 制造混乱 ----
# 1. 顶部几行空行和说明
ws["A1"] = "审计项目：ABC公司2024年年报审计"
ws.merge_cells("A1:F1")
ws["A1"].font = Font(bold=True, size=14)

ws["A2"] = "编制：张三    日期：2025-01-15"
ws.merge_cells("A2:F2")
ws["A2"].font = Font(italic=True, color="666666")

ws["A4"] = "重要提示：以下为抽凭清单，请按顺序翻凭证！"
ws.merge_cells("A4:F4")
ws["A4"].font = Font(bold=True, color="FF0000")
ws["A4"].alignment = Alignment(wrap_text=True)

# 2. 表头带合并单元格
ws["A6"] = "序号"
ws.merge_cells("A6:A7")
ws["B6"] = "凭证信息"
ws.merge_cells("B6:E6")
ws["F6"] = "备注"
ws.merge_cells("F6:F7")

ws["B7"] = "月份"
ws["C7"] = "凭证字"
ws["D7"] = "凭证号"
ws["E7"] = "金额"

# 3. 填数据 -- 制造各种脏格式
header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
for cell in ["A6","B6","C6","D6","E6","F6","A7","B7","C7","D7","E7","F7"]:
    ws[cell].fill = header_fill
    ws[cell].font = Font(bold=True)
    ws[cell].alignment = Alignment(horizontal="center", vertical="center")

# 数据行
sample_data = [
    # 正常行
    [1, "1月", "记", "001", 5000.00, ""],
    [2, "1月", "记", "002", 12300.50, "差旅费"],
    # 空行（模拟）
    [],
    [],
    [3, "1月", "转", "003", 250000.00, "固定资产采购"],
    # 合并单元格（模拟2行合并成1行）
    [4, "1月", "银", "004", 8888.88, "银行手续费"],
    # 金额文本格式（带逗号）
    [5, "2月", "记", "010", "12,500.00", "办公用品"],
    # 金额文本格式（带人民币符号）
    [6, "2月", "转", "011", "¥3,200.00", "咨询费"],
    # 空行
    [],
    [7, "2月", "记", "012", 6780.00, ""],
    # 日期格式不统一（有的"2025年3月"）
    [8, "2025年3月", "银", "020", 15000.00, "房租"],
    # 日期只有月份没有"月"字
    [9, "3", "记", "021", 4500.00, "快递费"],
    # 金额负数（借方红字）
    [10, "3月", "转", "022", -300.00, "冲回费用"],
    # 凭证号是文本（带0开头）
    [11, "4月", "记", "0035", 2200.00, "交通费"],
    # 整行合并过的假象
    [12, "4月", "记", "036", 9800.00, "招待费"],
    # 空行
    [],
    [],
    [13, "5月", "银", "045", 56000.00, "设备款"],
    # 金额带多余空格
    [14, "5月", "转", "046", " 6700 ", "维修费"],
    # 摘要为空
    [15, "5月", "记", "047", 3300.00, ""],
    # 月份有前缀0
    [16, "06月", "银", "055", 120000.00, "货款"],
    [17, "6月", "记", "056", 450.00, "报刊订阅"],
    # 凭证号数字超大
    [18, "7月", "转", "108", 8900.00, "培训费"],
    # 金额科学计数法风险
    [19, "7月", "记", "109", 78000.00, "软件采购"],
    [20, "8月", "银", "126", 3400.00, "水电费"],
]

row = 8
for item in sample_data:
    if not item:  # 空行
        row += 1
        continue
    for col_idx, val in enumerate(item, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
    row += 1

# 底部加个汇总
row += 1
ws.cell(row=row, column=1, value="合计：")
ws.cell(row=row, column=1).font = Font(bold=True)
ws.cell(row=row, column=5, value="=SUM(E8:E" + str(row-1) + ")")
ws.cell(row=row, column=5).font = Font(bold=True)

# 列宽
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 20

# 加个边框看起来更真
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

output_path = "/Users/zhangxin/Desktop/PQ练习_脏清单样本.xlsx"
wb.save(output_path)
print(f"已生成：{output_path}")
print("这是一个模拟Mentor给的烂格式清单，用来练Power Query清洗")
