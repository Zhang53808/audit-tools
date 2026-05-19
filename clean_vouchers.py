#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据清洗脚本 v2 —— 修复月份格式问题
"""

import openpyxl, re

print("请把脏清单Excel拖进终端，然后回车：")
path = input().strip().strip("'\"")

wb = openpyxl.load_workbook(path)
ws = wb.active

# 读取所有行，跳过空行
all_rows = []
for row in ws.iter_rows(min_row=1, values_only=True):
    all_rows.append(list(row))

# 找到真正的表头行
header_idx = None
for i, row in enumerate(all_rows):
    text = " ".join(str(v) for v in row if v)
    if "序号" in text and "月份" in text and "凭证" in text:
        header_idx = i
        break

if header_idx is None:
    # 换个查找方式
    for i, row in enumerate(all_rows):
        text = " ".join(str(v) for v in row if v)
        if "月" in text and "凭证" in text and "金额" in text:
            header_idx = i
            break

if header_idx is None:
    print("找不到表头行")
    exit(1)

headers = []
for j, v in enumerate(all_rows[header_idx]):
    if isinstance(v, str) and v.strip():
        headers.append((j, v.strip()))
    elif isinstance(v, (int, float)):
        headers.append((j, str(v)))

# 找出各列的位置
col_index = {}  # 列名 -> 列号
for j, h in headers:
    if "月" in h:
        col_index["month"] = j
    elif "凭证字" in h:
        col_index["type"] = j
    elif "凭证号" in h:
        col_index["num"] = j
    elif "金额" in h:
        col_index["amount"] = j
    elif "备注" in h or "摘要" in h:
        col_index["note"] = j

print(f"识别到的列：{col_index}")

# 清洗数据
data = all_rows[header_idx + 1:]
cleaned = []

def clean_month(val):
    """把各种月份的写法统一成"X月" """
    if val is None:
        return ""
    s = str(val).strip()
    if not s:
        return ""
    # 优先匹配带"年"的："2025年3月" -> "3月"
    m = re.search(r'年\s*(\d+)\s*月', s)
    if m:
        num = int(m.group(1))
        if 1 <= num <= 12:
            return f"{num}月"
    # 没年，直接找数字+月："06月"、"3月" -> "6月"、"3月"
    m = re.search(r'(\d+)\s*月', s)
    if m:
        num = int(m.group(1))
        return f"{num}月"
    # 只有数字没有月："3" -> "3月"
    m = re.search(r'^(\d+)$', s)
    if m:
        num = int(m.group(1))
        if 1 <= num <= 12:
            return f"{num}月"
    return s

def clean_amount(val):
    """清洗金额，去掉¥和逗号"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    s = s.replace("¥", "").replace(",", "").replace(" ", "")
    try:
        return float(s)
    except:
        return 0

for row in data:
    # 跳过空行
    if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
        continue
    # 跳过合计行
    if row[0] and "合计" in str(row[0]):
        continue
    
    new_row = list(row)  # 复制
    
    if "month" in col_index:
        j = col_index["month"]
        new_row[j] = clean_month(new_row[j])
    
    if "amount" in col_index:
        j = col_index["amount"]
        new_row[j] = clean_amount(new_row[j])
    
    # 跳过序号为空的行
    if row[0] is None or (isinstance(row[0], str) and row[0].strip() == ""):
        continue
    
    cleaned.append(new_row)

# 输出
out_path = path.replace(".xlsx", "_已清洗.xlsx")
out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = "清洗结果"

# 表头
for j, h in headers:
    out_ws.cell(row=1, column=j+1, value=h)

# 数据
for i, row in enumerate(cleaned, 2):
    for j in range(len(row)):
        out_ws.cell(row=i, column=j+1, value=row[j])

out_wb.save(out_path)
print(f"清洗完成！结果：{out_path}")
print(f"有效数据：{len(cleaned)} 行")

# 预览
print("\n预览（月份列 + 金额列）：")
for row in cleaned[:5]:
    m = row[col_index["month"]] if "month" in col_index else "?"
    a = row[col_index["amount"]] if "amount" in col_index else "?"
    print(f"  月份：{m}    金额：{a}")
