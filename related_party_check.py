#!/usr/bin/env python3
"""
关联方识别 - 12维度交叉比对引擎
=================================
逆向狗(nigo)方法论复现 + 卓朗科技案复现

用法：
    python related_party_check.py                         # 使用内置模拟数据
    python related_party_check.py 输入.xlsx                # 从Excel加载企业名单

输出：
    关联方核查结果.xlsx（12维矩阵 + 三色风险等级）

依赖安装：
    pip install pandas openpyxl thefuzz python-Levenshtein
"""

import os
import sys
import re
import json
import argparse
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from collections import defaultdict

import pandas as pd
from thefuzz import fuzz

# ============================================================
# 地址清洗（从 address_verification 复制核心函数，避免循环导入）
# ============================================================
def _clean_address(addr: str) -> str:
    """清洗地址：去行政前缀、统一标点、去空白。"""
    if not isinstance(addr, str) or not addr.strip():
        return ""
    result = addr.strip()
    result = result.replace("（", "(").replace("）", ")")
    result = result.replace("，", ",").replace("。", ".")
    result = result.replace("；", ";").replace("：", ":")
    result = result.replace("－", "-").replace("—", "-")
    result = result.replace("’", "'").replace("‘", "'")
    result = result.replace("“", '"').replace("”", '"')
    result = re.sub(r'\s+', '', result)
    result = re.sub(
        r'^(北京市|天津市|上海市|重庆市|'
        r'河北省|山西省|辽宁省|吉林省|黑龙江省|'
        r'江苏省|浙江省|安徽省|福建省|江西省|山东省|'
        r'河南省|湖北省|湖南省|广东省|海南省|'
        r'四川省|贵州省|云南省|陕西省|甘肃省|青海省|'
        r'台湾省|'
        r'内蒙古自治区|广西壮族自治区|西藏自治区|宁夏回族自治区|新疆维吾尔自治区|'
        r'香港特别行政区|澳门特别行政区)',
        '', result
    )
    abbreviations = {
        "高新技术产业开发区": "高新区",
        "经济技术开发区": "开发区",
        "化学工业区": "化工区",
        "保税港区": "保税区",
    }
    for full, abbr in abbreviations.items():
        result = result.replace(full, abbr)
    prev = None
    while prev != result:
        prev = result
        result = re.sub(
            r'^([一-龥]{2,6}(?:市|区|县|州|盟|旗|自治州|自治县|自治旗))',
            '', result
        )
    return result


def _address_similarity(addr1: str, addr2: str) -> float:
    """地址多策略相似度（复用地址核查逻辑）。"""
    s1 = _clean_address(addr1)
    s2 = _clean_address(addr2)
    if not s1 or not s2:
        return 0.0
    return max(fuzz.ratio(s1, s2), fuzz.token_sort_ratio(s1, s2), fuzz.partial_ratio(s1, s2))


def _name_similarity(name1: str, name2: str) -> float:
    """企业名称相似度（去除通用后缀后比较）。"""
    if not name1 or not name2:
        return 0.0
    # 去除通用公司后缀，避免"XX有限公司"和"YY有限公司"被后缀拉高相似度
    suffixes = [
        "股份有限公司", "有限责任公司", "有限公司", "股份公司",
        "(特殊普通合伙)", "(普通合伙)", "(有限合伙)",
    ]
    n1, n2 = name1, name2
    for s in suffixes:
        n1 = n1.replace(s, "")
        n2 = n2.replace(s, "")
    return max(fuzz.ratio(n1, n2), fuzz.token_sort_ratio(n1, n2))


def _extract_names(items: List[str]) -> set:
    """从 '张某某(30%)' 格式中提取纯人名/企业名集合。"""
    names = set()
    for item in items:
        if isinstance(item, str):
            # 去掉括号里的比例/职务
            name = re.sub(r'\(.*?\)', '', item).strip()
            if name:
                names.add(name)
    return names


# ============================================================
# 模拟数据：卓朗科技案 (8家隐性关联方 + 1个干扰项)
# ============================================================
def _build_fixture() -> Tuple[dict, List[dict]]:
    """
    构建卓朗科技案模拟数据集。

    返回: (审计方数据, [目标企业列表])
    """
    audit = {
        "name": "卓朗科技股份有限公司",
        "shareholders": ["张某某(30%)", "李某某(25%)", "天津卓朗控股有限公司(45%)"],
        "actual_controller": "张某某",
        "executives": ["张某某(董事长)", "王某某(总经理)", "赵某(监事)"],
        "legal_rep": "张某某",
        "address": "天津市滨海新区经济技术开发区信环西路20号",
        "phone": "022-5988****",
        "email": "info@zhuolang.com",
    }

    targets = [
        {
            # 1. D08信号: 电话138****5678 = 审计方监事赵某的个人号码
            "name": "温岭乾民贸易有限公司",
            "shareholders": ["王某(60%)", "李某(40%)"],
            "actual_controller": "王某",
            "executives": ["王某(执行董事)", "李某(经理)", "赵某(监事)"],
            "legal_rep": "王某",
            "historical_legal_reps": [],
            "address": "天津市滨海新区塘沽街道新港路100号",
            "phone": "138****5678",
            "email": "wenling@qianmin-trade.cn",
            "change_dates": [],
            "insured_employees": 5,
            "transaction_amount": 5000000,
        },
        {
            # 2. D02+D04信号: 股东梁某也在#3 + 高管王某某与审计方重叠
            "name": "天津事达科技有限公司",
            "shareholders": ["梁某(50%)", "陈某(50%)"],
            "actual_controller": "梁某",
            "executives": ["梁某(执行董事)", "王某某(经理)"],
            "legal_rep": "梁某",
            "historical_legal_reps": [],
            "address": "天津市南开区鞍山西道200号",
            "phone": "022-8790****",
            "email": "shida@tj-shida.cn",
            "change_dates": [],
            "insured_employees": 8,
            "transaction_amount": 8500000,
        },
        {
            # 3. D02信号: 股东梁某与#2重叠
            "name": "事达信息咨询有限公司",
            "shareholders": ["梁某(40%)", "张某(30%)", "赵某(30%)"],
            "actual_controller": "梁某",
            "executives": ["梁某(执行董事)", "张某(经理)"],
            "legal_rep": "梁某",
            "historical_legal_reps": [],
            "address": "天津市南开区鞍山西道210号",
            "phone": "022-8791****",
            "email": "shida_info@tj-shida.cn",
            "change_dates": [],
            "insured_employees": 3,
            "transaction_amount": 3200000,
        },
        {
            # 4. D10+D07信号: 名字含"卓" + 地址靠近审计方
            "name": "滨海新区卓创商贸有限公司",
            "shareholders": ["刘某(100%)"],
            "actual_controller": "刘某",
            "executives": ["刘某(执行董事)", "孙某(监事)"],
            "legal_rep": "刘某",
            "historical_legal_reps": [],
            "address": "天津市滨海新区经济技术开发区信环西路18号",
            "phone": "022-5900****",
            "email": "zhuochuang@bh-zhuoc.cn",
            "change_dates": [],
            "insured_employees": 2,
            "transaction_amount": 1800000,
        },
        {
            # 5. D10信号: 名字含"朗" + D08信号: 电话与#1相同
            "name": "天津朗信科技有限公司",
            "shareholders": ["周某(70%)", "吴某(30%)"],
            "actual_controller": "周某",
            "executives": ["周某(执行董事)", "吴某(监事)"],
            "legal_rep": "周某",
            "historical_legal_reps": [],
            "address": "天津市西青区中北镇星光路50号",
            "phone": "138****5678",
            "email": "langxin@langxin-tech.cn",
            "change_dates": [],
            "insured_employees": 4,
            "transaction_amount": 4200000,
        },
        {
            # 6. D10+D06信号: 名字含"乾民" + 前法人赵某=审计方监事
            "name": "乾民信息技术有限公司",
            "shareholders": ["陈某(55%)", "黄某(45%)"],
            "actual_controller": "陈某",
            "executives": ["陈某(执行董事)", "黄某(经理)"],
            "legal_rep": "陈某",
            "historical_legal_reps": ["赵某"],   # ← 关键：前法人
            "address": "天津市河东区十一经路80号",
            "phone": "022-2410****",
            "email": "qianmin@qm-info.cn",
            "change_dates": [],
            "insured_employees": 1,
            "transaction_amount": 1200000,
        },
        {
            # 7. D03信号: 实际控制人张某某 = 审计方大股东/实际控制人
            "name": "北辰区明锐达科技有限公司",
            "shareholders": ["张某某(80%)", "林某(20%)"],
            "actual_controller": "张某某",
            "executives": ["张某某(执行董事)", "林某(监事)"],
            "legal_rep": "张某某",
            "historical_legal_reps": [],
            "address": "天津市北辰区京津公路300号",
            "phone": "022-2680****",
            "email": "mingruida@mingruida-bc.cn",
            "change_dates": [],
            "insured_employees": 0,
            "transaction_amount": 9500000,
        },
        {
            # 8. D10+D01信号: 名字含"卓" + 审计方通过控股公司间接持股
            "name": "天津卓研咨询有限公司",
            "shareholders": ["天津卓朗控股有限公司(51%)", "侯某(49%)"],
            "actual_controller": "张某某",
            "executives": ["侯某(执行董事)", "侯某(经理)"],
            "legal_rep": "侯某",
            "historical_legal_reps": [],
            "address": "天津市和平区南京路100号",
            "phone": "022-2330****",
            "email": "zhuoyan@zhuoyan-tj.cn",
            "change_dates": [],
            "insured_employees": 6,
            "transaction_amount": 7200000,
        },
        {
            # 9. 干扰项：万科 — 应12维度全✗
            "name": "万科企业股份有限公司",
            "shareholders": ["深圳市地铁集团有限公司(27.91%)", "香港中央结算有限公司(7.34%)", "郁某(0.08%)"],
            "actual_controller": "深圳市地铁集团有限公司",
            "executives": ["郁某(董事长)", "祝某(总裁)", "韩某(执行副总裁)"],
            "legal_rep": "郁某",
            "historical_legal_reps": [],
            "address": "深圳市盐田区大梅沙环梅路33号万科中心",
            "phone": "0755-2560****",
            "email": "ir@vanke.com",
            "change_dates": [],
            "insured_employees": 10000,
            "transaction_amount": 500000000,
        },
    ]

    return audit, targets


# ============================================================
# 12维度检测函数
# ============================================================
CheckResult = Tuple[bool, str]  # (是否触发, 证据描述)


def check_D01_equity(audit: dict, target: dict) -> CheckResult:
    """
    D01 股权穿透: 审计方与客户有股权关系。
    三种情况：
      1. 共同股东（审计方股东 ∩ 客户股东 ≠ ∅）
      2. 客户是审计方的股东（母公司）
      3. 审计方是客户的股东（子公司/投资企业）
    """
    audit_holders = _extract_names(audit.get("shareholders", []))
    target_holders = _extract_names(target.get("shareholders", []))

    # 情况1: 共同股东
    common = audit_holders & target_holders
    if common:
        return True, f"共同股东: {', '.join(sorted(common))}"

    # 情况2: 客户本身就是审计方的股东（母公司）
    audit_name = audit.get("name", "")
    target_name = target.get("name", "")
    # 去后缀比较
    for holder_name in audit_holders:
        sim = _name_similarity(target_name, holder_name)
        if sim >= 70:
            return True, f"客户「{target_name}」是审计方股东「{holder_name}」(相似度{sim}%)"

    # 情况3: 审计方名称出现在客户股东列表中
    for holder_name in target_holders:
        sim = _name_similarity(audit_name, holder_name)
        if sim >= 70:
            return True, f"审计方「{audit_name}」是客户股东「{holder_name}」(相似度{sim}%)"

    return False, ""


def check_D02_common_shareholders(targets: List[dict]) -> Dict[str, CheckResult]:
    """
    D02 共同股东: 多家客户背后是否有同一批人。
    算法: 所有客户两两比较股东集合的交集，返回每个企业是否有共同股东。
    """
    results: Dict[str, CheckResult] = {}
    n = len(targets)
    triggered = set()

    for i in range(n):
        for j in range(i + 1, n):
            si = _extract_names(targets[i].get("shareholders", []))
            sj = _extract_names(targets[j].get("shareholders", []))
            common = si & sj
            if common:
                triggered.add(targets[i]["name"])
                triggered.add(targets[j]["name"])
                names_str = ', '.join(sorted(common))
                # 存到结果的证据里
                for t in [targets[i], targets[j]]:
                    name = t["name"]
                    old_evidence = results.get(name, (False, ""))[1]
                    other = targets[i]["name"] if targets[j]["name"] == name else targets[j]["name"]
                    new_ev = f"与「{other}」共同股东: {names_str}"
                    combined = "; ".join([e for e in [old_evidence, new_ev] if e])
                    results[name] = (True, combined)

    # 未触发的企业也填入 False
    for t in targets:
        if t["name"] not in results:
            results[t["name"]] = (False, "")

    return results


def check_D03_actual_controller(audit: dict, target: dict) -> CheckResult:
    """
    D03 实际控制人: 最终受益人是否相同。
    """
    audit_ctrl = audit.get("actual_controller", "")
    target_ctrl = target.get("actual_controller", "")
    if audit_ctrl and target_ctrl and audit_ctrl == target_ctrl:
        return True, f"实际控制人相同: {audit_ctrl}"
    return False, ""


def check_D04_executive_overlap(audit: dict, target: dict) -> CheckResult:
    """
    D04 关键管理人员重叠: 同一人在两边当高管。
    """
    audit_execs = _extract_names(audit.get("executives", []))
    target_execs = _extract_names(target.get("executives", []))
    common = audit_execs & target_execs
    if common:
        return True, f"高管重叠: {', '.join(sorted(common))}"
    return False, ""


def check_D05_legal_rep_cross(targets: List[dict]) -> Dict[str, CheckResult]:
    """
    D05 法定代表人交叉: 同一人当多家企业法人。
    """
    results: Dict[str, CheckResult] = {}
    legal_map: Dict[str, List[str]] = defaultdict(list)

    for t in targets:
        lr = t.get("legal_rep", "")
        if lr:
            legal_map[lr].append(t["name"])

    for lr, companies in legal_map.items():
        if len(companies) > 1:
            for name in companies:
                others = [c for c in companies if c != name]
                results[name] = (
                    True,
                    f"与「{'」「'.join(others)}」共用法人: {lr}"
                )

    # 未触发的
    for t in targets:
        if t["name"] not in results:
            results[t["name"]] = (False, "")

    return results


def check_D06_legal_rep_trajectory(audit: dict, target: dict) -> CheckResult:
    """
    D06 法人变更轨迹: 客户前任法人是否 = 审计方高管/监事。
    """
    audit_people = _extract_names(audit.get("executives", []))
    audit_people.add(audit.get("legal_rep", ""))
    target_historical = _extract_names(target.get("historical_legal_reps", []))

    common = audit_people & target_historical
    if common:
        return True, f"前法人「{', '.join(sorted(common))}」现为审计方人员"
    return False, ""


def check_D07_same_address(audit: dict, target: dict) -> CheckResult:
    """
    D07 同址经营: 注册地址相同或高度相似（≥70%）。
    """
    a1 = audit.get("address", "")
    a2 = target.get("address", "")
    sim = _address_similarity(a1, a2)
    if sim >= 70:
        return True, f"地址相似度 {sim}%: 「{a1}」vs「{a2}」"
    return False, ""


def check_D08_shared_contact(targets: List[dict]) -> Dict[str, CheckResult]:
    """
    D08 联系方式共用: 电话或邮箱域名相同。
    """
    results: Dict[str, CheckResult] = {}
    phone_map: Dict[str, List[str]] = defaultdict(list)
    email_domain_map: Dict[str, List[str]] = defaultdict(list)

    for t in targets:
        phone = t.get("phone", "")
        email = t.get("email", "")
        if phone:
            phone_map[phone].append(t["name"])
        if email and "@" in email:
            domain = email.split("@")[1].strip().lower()
            # 排除公共邮箱域名（个人邮箱，不代表公司关联）
            public_domains = {
                "qq.com", "163.com", "126.com", "139.com", "sina.com",
                "sohu.com", "yeah.net", "gmail.com", "outlook.com",
                "hotmail.com", "foxmail.com", "189.cn", "wo.cn",
                "aliyun.com", "vip.qq.com", "vip.163.com",
            }
            if domain not in public_domains:
                email_domain_map[domain].append(t["name"])

    # 共用电话
    for phone, companies in phone_map.items():
        if len(companies) > 1:
            for name in companies:
                others = [c for c in companies if c != name]
                results[name] = (
                    True,
                    f"与「{'」「'.join(others)}」共用电话: {phone}"
                )

    # 共用邮箱域名
    for domain, companies in email_domain_map.items():
        if len(companies) > 1:
            for name in companies:
                existing = results.get(name)
                others = [c for c in companies if c != name]
                evidence = f"与「{'」「'.join(others)}」共用邮箱域名: @{domain}"
                if existing:
                    results[name] = (True, f"{existing[1]}; {evidence}")
                else:
                    results[name] = (True, evidence)

    # 未触发
    for t in targets:
        if t["name"] not in results:
            results[t["name"]] = (False, "")

    return results


def check_D09_change_window(audit: dict, target: dict, transaction_date: Optional[str] = None) -> CheckResult:
    """
    D09 变更时间窗口: 交易前后N天内有工商变更。
    模拟数据中无变更日期，暂跳过，留接口。
    """
    return False, ""


def check_D10_name_similarity(audit: dict, target: dict) -> CheckResult:
    """
    D10 企业名称相似: 名字长得像（≥50%）或含有关键共用字。

    策略:
      1. 去后缀后的 fuzz 相似度 ≥50%
      2. 审计方核心2字词出现在目标名称中
      3. 共同含有关键单字（非通用字）
    """
    a_name = audit.get("name", "")
    t_name = target.get("name", "")
    sim = _name_similarity(a_name, t_name)
    if sim >= 50:
        return True, f"名称相似度 {sim}%: 「{a_name}」vs「{t_name}」"

    # 子策略1：名称中是否含有审计方的核心关键词（2字窗口）
    # 注意：必须在去后缀后的名称上扫描，避免"限公""份有"等后缀片段误匹配
    a_clean = a_name
    t_clean = t_name
    for s in ["股份有限公司", "有限责任公司", "有限公司", "股份公司"]:
        a_clean = a_clean.replace(s, "")
        t_clean = t_clean.replace(s, "")

    blocked = {"有限", "公司", "股份", "责任", "科技", "企业", "集团", "实业", "发展", "限公", "份有", "司股"}
    for i in range(len(a_clean) - 1):
        chunk = a_clean[i:i+2]
        if chunk not in blocked and chunk in t_clean:
            return True, f"名称含共同片段「{chunk}」"

    # 子策略2：单字共现（对2字公司名核心字有效，如"卓""朗"）
    a_stripped = a_name
    for s in ["股份有限公司", "有限责任公司", "有限公司", "股份公司"]:
        a_stripped = a_stripped.replace(s, "")
    t_stripped = t_name
    for s in ["股份有限公司", "有限责任公司", "有限公司", "股份公司"]:
        t_stripped = t_stripped.replace(s, "")

    common_chars = set(a_stripped) & set(t_stripped)
    trivial = {"科", "技", "信", "息", "商", "贸", "工", "实", "业", "发", "展"}
    meaningful = common_chars - trivial
    if meaningful:
        return True, f"名称含共同特征字「{'」「'.join(sorted(meaningful))}」"

    return False, ""


def check_D11_insured_anomaly(target: dict, threshold: int = 5000000) -> CheckResult:
    """
    D11 参保人数异常: 0人参保但交易额巨大。
    """
    insured = target.get("insured_employees", 0)
    amount = target.get("transaction_amount", 0)
    if insured == 0 and amount >= threshold:
        return True, f"0人参保但交易额达 {amount/10000:.0f}万"
    return False, ""


def check_D12_supplement_personnel(
    target: dict,
    personnel_list: Optional[List[str]] = None,
) -> CheckResult:
    """
    D12 补充人员关联: 补充名单匹配客户高管/股东。
    """
    if not personnel_list:
        return False, ""

    target_people = _extract_names(target.get("executives", []))
    target_people |= _extract_names(target.get("shareholders", []))
    target_people.add(target.get("legal_rep", ""))

    personnel_set = set(p.strip() for p in personnel_list if p.strip())
    matches = target_people & personnel_set

    if matches:
        return True, f"补充名单匹配: {', '.join(sorted(matches))}"
    return False, ""


# ============================================================
# 风险评分
# ============================================================
def calculate_risk(flag_count: int, has_high_weight: bool) -> Tuple[str, str]:
    """
    映射到风险等级。

    基础规则:
      0     → 低风险
      1-2   → 中风险
      ≥3    → 高风险

    权重升级: D01/D03/D06 任一触发 → 升一档
    """
    if flag_count == 0:
        level = "低风险"
    elif flag_count <= 2:
        level = "中风险"
    else:
        level = "高风险"

    if has_high_weight and level == "低风险":
        level = "中风险"
    elif has_high_weight and level == "中风险":
        level = "高风险"

    icon = {"低风险": "🟢", "中风险": "🟡", "高风险": "🔴"}
    return level, f"{icon.get(level, '')} {level}"


# ============================================================
# 主流程
# ============================================================
def run_check(audit: dict, targets: List[dict],
              personnel_list: Optional[List[str]] = None,
              output_file: Optional[str] = None) -> pd.DataFrame:
    """
    跑12维度全量比对。

    Args:
        audit: 审计方数据
        targets: 客户/供应商列表
        personnel_list: 补充人员名单（用于D12）
        output_file: 输出Excel路径（可选）

    Returns:
        结果DataFrame
    """
    print(f"🔍 被审计单位: {audit['name']}")
    print(f"📋 待核查企业: {len(targets)} 家\n")

    dim_labels = [
        "股权穿透", "共同股东", "实际控制人", "高管重叠",
        "法人交叉", "法人变更", "同址经营", "联系方式共用",
        "变更时间窗口", "名称相似", "参保异常", "人员关联",
    ]
    DIM_KEYS = {  # 中文名 → 内部键名映射
        "股权穿透": "D01", "共同股东": "D02", "实际控制人": "D03",
        "高管重叠": "D04", "法人交叉": "D05", "法人变更": "D06",
        "同址经营": "D07", "联系方式共用": "D08", "变更时间窗口": "D09",
        "名称相似": "D10", "参保异常": "D11", "人员关联": "D12",
    }
    HIGH_WEIGHT_DIMS = {"股权穿透", "实际控制人", "法人变更"}  # 高权重维度

    # ---- 第一步：逐条跑独立维度 ----
    rows = []
    for t in targets:
        name = t["name"]
        flags = {}

        flags["股权穿透"] = check_D01_equity(audit, t)
        flags["实际控制人"] = check_D03_actual_controller(audit, t)
        flags["高管重叠"] = check_D04_executive_overlap(audit, t)
        flags["法人变更"] = check_D06_legal_rep_trajectory(audit, t)
        flags["同址经营"] = check_D07_same_address(audit, t)
        flags["变更时间窗口"] = check_D09_change_window(audit, t)
        flags["名称相似"] = check_D10_name_similarity(audit, t)
        flags["参保异常"] = check_D11_insured_anomaly(t)
        flags["人员关联"] = check_D12_supplement_personnel(t, personnel_list)

        rows.append({"name": name, "_target": t, "_flags": flags})

    # ---- 第二步：跑跨企业维度 ----
    d02_results = check_D02_common_shareholders(targets)
    d05_results = check_D05_legal_rep_cross(targets)
    d08_results = check_D08_shared_contact(targets)

    for row in rows:
        name = row["name"]
        row["_flags"]["共同股东"] = d02_results.get(name, (False, ""))
        row["_flags"]["法人交叉"] = d05_results.get(name, (False, ""))
        row["_flags"]["联系方式共用"] = d08_results.get(name, (False, ""))

    # ---- 第三步：汇总 + 风险评分 + 数据可用性标记 ----
    data = []
    for row in rows:
        name = row["name"]
        t = row["_target"]
        flags = row["_flags"]

        # 检查每个维度的数据是否可用
        data_available = {}
        data_available["股权穿透"] = bool(t.get("shareholders")) and bool(audit.get("shareholders"))
        data_available["共同股东"] = bool(t.get("shareholders"))
        data_available["实际控制人"] = bool(t.get("actual_controller")) and bool(audit.get("actual_controller"))
        data_available["高管重叠"] = bool(t.get("executives")) and bool(audit.get("executives"))
        data_available["法人交叉"] = bool(t.get("legal_rep"))
        data_available["法人变更"] = bool(t.get("historical_legal_reps")) and bool(audit.get("executives"))
        data_available["同址经营"] = bool(t.get("address")) and bool(audit.get("address"))
        data_available["联系方式共用"] = bool(t.get("phone")) or bool(t.get("email"))
        data_available["变更时间窗口"] = bool(t.get("change_dates"))  # 通常为 False
        data_available["名称相似"] = True   # 始终可用，只需企业名
        data_available["参保异常"] = "insured_employees" in t and "transaction_amount" in t
        data_available["人员关联"] = bool(personnel_list)

        # 只计算数据可用的维度中的触发数
        flag_count = 0
        for dim_cn in dim_labels:
            triggered = flags.get(dim_cn, (False,))[0]
            if triggered and data_available.get(dim_cn, False):
                flag_count += 1
            # 如果维度的数据不可用，不计数（无法判断）
        has_high = any(
            flags.get(d, (False,))[0] and data_available.get(d, False)
            for d in HIGH_WEIGHT_DIMS
        )
        risk_level, risk_label = calculate_risk(flag_count, has_high)

        record = {"企业名称": name}
        for dim_cn in dim_labels:
            triggered, evidence = flags.get(dim_cn, (False, ""))
            if not data_available.get(dim_cn, False):
                record[dim_cn] = "—"     # 数据不全，无法判断
            elif triggered:
                record[dim_cn] = evidence
            else:
                record[dim_cn] = "✗"
        record["异常维度数"] = flag_count
        record["风险等级"] = risk_label
        data.append(record)

        icon = "🔴" if risk_level == "高风险" else ("🟡" if risk_level == "中风险" else "🟢")
        triggered_dims = [
            dim_cn for dim_cn in dim_labels
            if flags.get(dim_cn, (False,))[0] and data_available.get(dim_cn, False)
        ]
        dims_str = ", ".join(triggered_dims) if triggered_dims else "无"
        print(f"  {icon} {name}: {dims_str}")

    df = pd.DataFrame(data)

    # ---- 统计 ----
    stats = df["风险等级"].apply(lambda x: x.split()[-1]).value_counts()
    # 统计数据可用性
    unavailable_dims = []
    for dim_cn in dim_labels:
        if all(row[dim_cn] == "—" for row in data if row[dim_cn] != "✗" and row[dim_cn] != "—"):
            # 全部不可用
            pass
        has_unavailable = sum(1 for row in data if row[dim_cn] == "—")
    print(f"\n{'='*50}")
    print(f"📊 核查完成")
    print(f"   总数: {len(targets)}")
    print(f"   🟢 低风险: {stats.get('低风险', 0)}")
    print(f"   🟡 中风险: {stats.get('中风险', 0)}")
    print(f"   🔴 高风险: {stats.get('高风险', 0)}")
    # 标出哪些维度数据不全
    partial_dims = [
        dim_cn for dim_cn in dim_labels
        if any(row[dim_cn] == "—" for row in data)
    ]
    if partial_dims:
        print(f"\n⚠️ 以下维度数据不全，标注为「—」：")
        for d in partial_dims:
            missing = sum(1 for row in data if row[d] == "—")
            print(f"   {d}: {missing}/{len(targets)} 家企业缺数据")

    # ---- 输出Excel ----
    if output_file is None:
        output_file = "关联方核查结果.xlsx"

    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='关联方核查')

    ws = writer.sheets['关联方核查']

    # 列宽
    col_widths = {"企业名称": 28}
    for d in dim_labels:
        col_widths[d] = 32       # 证据文字需要宽列
    col_widths["异常维度数"] = 10
    col_widths["风险等级"] = 14

    for col_idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 10)

    # 三色条件格式
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_font = Font(color="006100")
    yellow_font = Font(color="9C6500")
    red_font = Font(color="9C0006")

    risk_col = len(df.columns)
    risk_letter = get_column_letter(risk_col)

    for row_idx in range(2, len(df) + 2):
        cell = ws[f"{risk_letter}{row_idx}"]
        if cell.value and "低风险" in str(cell.value):
            cell.fill = green_fill
            cell.font = green_font
        elif cell.value and "中风险" in str(cell.value):
            cell.fill = yellow_fill
            cell.font = yellow_font
        elif cell.value and "高风险" in str(cell.value):
            cell.fill = red_fill
            cell.font = red_font

    # ✓ 列绿色、✗ 不染色
    for dim in dim_labels:
        col_idx = df.columns.get_loc(dim) + 1
        col_letter = get_column_letter(col_idx)
        for row_idx in range(2, len(df) + 2):
            cell = ws[f"{col_letter}{row_idx}"]
            if cell.value and cell.value not in ("✗", "—"):
                cell.fill = green_fill
                cell.font = green_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"

    writer.close()
    print(f"   报告已保存: {output_file}")

    return df


# ============================================================
# 数据加载
# ============================================================
def load_targets_from_excel(filepath: str) -> List[dict]:
    """
    从Excel加载目标企业。

    期望列（英或中均可）:
      企业名称 / name
      股东 / shareholders
      实际控制人 / actual_controller
      高管 / executives
      法定代表人 / legal_rep
      注册地址 / address
      电话 / phone
      邮箱 / email
      参保人数 / insured_employees
      交易金额 / transaction_amount
      历史法人 / historical_legal_reps
    """
    df = pd.read_excel(filepath)

    col_map = {
        "name": ["企业名称", "公司名称", "name", "名称"],
        "shareholders": ["股东", "股东信息", "shareholders"],
        "actual_controller": ["实际控制人", "最终受益人", "actual_controller"],
        "executives": ["高管", "关键管理人员", "executives"],
        "legal_rep": ["法定代表人", "法人", "legal_rep"],
        "address": ["注册地址", "地址", "address"],
        "phone": ["电话", "联系电话", "phone"],
        "email": ["邮箱", "email"],
        "insured_employees": ["参保人数", "insured_employees"],
        "transaction_amount": ["交易金额", "transaction_amount"],
        "historical_legal_reps": ["历史法人", "historical_legal_reps"],
    }

    targets = []
    for _, row in df.iterrows():
        t = {}
        for key, candidates in col_map.items():
            for c in candidates:
                if c in df.columns:
                    val = row[c]
                    if pd.notna(val):
                        if key in ("shareholders", "executives", "historical_legal_reps"):
                            # 分号或换行分隔的列表
                            raw = str(val)
                            t[key] = [x.strip() for x in re.split(r'[;\n]', raw) if x.strip()]
                        elif key in ("insured_employees", "transaction_amount"):
                            t[key] = int(float(val))
                        else:
                            t[key] = str(val)
                    break
        # 必须有名称为空则跳过
        if not t.get("name"):
            continue
        targets.append(t)

    return targets


# ============================================================
# 入口
# ============================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="关联方识别 - 12维度交叉比对引擎"
    )
    parser.add_argument(
        "input", nargs="?",
        help="输入Excel路径（含企业名单和工商数据）。不提供则使用内置模拟数据。"
    )
    parser.add_argument(
        "-o", "--output",
        default="关联方核查结果.xlsx",
        help="输出Excel路径"
    )
    parser.add_argument(
        "--audit-file",
        help="审计方数据Excel（可选，不提供则从输入Excel中找或使用内置数据）"
    )
    parser.add_argument(
        "--personnel-file",
        help="补充人员名单Excel（可选，用于D12）"
    )
    args = parser.parse_args()

    # 加载审计方数据
    audit = None
    if args.audit_file:
        entries = load_targets_from_excel(args.audit_file)
        if entries:
            audit = entries[0]
            print(f"📌 审计方: {audit['name']} (从文件加载)")

    # 加载人员补充名单
    personnel = None
    if args.personnel_file:
        pdf = pd.read_excel(args.personnel_file)
        # 找"姓名"列
        name_cols = ["姓名", "name", "名称", "人员"]
        for nc in name_cols:
            if nc in pdf.columns:
                personnel = pdf[nc].dropna().astype(str).tolist()
                break
        print(f"👥 补充人员名单: {len(personnel) if personnel else 0} 人")

    # 加载目标企业
    if args.input:
        targets = load_targets_from_excel(args.input)
        print(f"📋 从文件加载 {len(targets)} 家企业")
    else:
        if audit is None:
            audit, targets = _build_fixture()
            print("📦 使用内置模拟数据（卓朗科技案）")
        else:
            # 有审计方但没有目标企业文件 — 也降级到模拟数据
            _, targets = _build_fixture()
            print("📦 使用内置模拟数据（卓朗科技案）")

    if not targets:
        print("❌ 未找到任何待核查企业，退出。")
        sys.exit(1)

    df = run_check(audit, targets, personnel, output_file=args.output)
