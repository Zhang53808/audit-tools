#!/usr/bin/env python3
"""
函证地址核查 - 三层过滤自动化脚本（异常初筛器）
================================================
逆向狗(nigo)方法论复现 + 三色分级 + LLM内容评估

用法：
    1. 准备Excel（公司名称 | 发函地址 | 工商注册地址）
    2. 配置腾讯地图API Key
    3. python address_verification.py 输入.xlsx --map-key KEY [--llm-key KEY]

输出：
    核查结果.xlsx（绿/黄/红三色标记 + 风险评分 + 判定理由）

依赖安装：
    pip install pandas openpyxl thefuzz requests python-Levenshtein python-dotenv tqdm
    （LLM评估用DeepSeek API，无需额外SDK，直接用requests调OpenAI兼容接口）

配置方式（三选一，优先级从高到低）：
    1. 命令行参数: --map-key KEY --llm-key KEY
    2. 环境变量:   export TENCENT_MAP_KEY=xxx  export DEEPSEEK_API_KEY=xxx
    3. .env 文件:  在同目录创建 .env 文件（参考 .env.example）
"""

import os
import sys
import re
import json
import math
import argparse
from pathlib import Path
from typing import Optional, Tuple

import pandas as pd
import requests
from thefuzz import fuzz
from tqdm import tqdm

# 加载 .env 文件（如果存在）
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))
except ImportError:
    pass


# ============================================================
# 配置区（使用前填写，或通过命令行参数传入）
# ============================================================
CONFIG = {
    # 腾讯地图API Key（免费注册：https://lbs.qq.com/）
    # 优先级：命令行 --map-key > 环境变量 TENCENT_MAP_KEY > .env 文件
    "tencent_map_key": os.getenv("TENCENT_MAP_KEY", ""),

    # DeepSeek API Key（用于第2层LLM内容评估）
    # 优先级：命令行 --llm-key > 环境变量 DEEPSEEK_API_KEY > .env 文件
    "llm_api_key": os.getenv("DEEPSEEK_API_KEY", ""),

    # LLM配置（DeepSeek OpenAI兼容接口，也可替换为其他兼容服务）
    "llm_base_url": os.getenv("LLM_BASE_URL", "https://api.deepseek.com"),
    "llm_model": os.getenv("LLM_MODEL", "deepseek-chat"),

    # AnySearch CLI路径（开源用户需自行配置）
    "anysearch_cli": os.getenv("ANYSEARCH_CLI", ""),

    # ---- 第1层：文本相似度阈值 ----
    "sim_high": 85,   # ≥此值直接绿色通过
    "sim_mid": 70,    # <此值即使距离近也需关注

    # ---- 第1.5层：地理距离阈值（米） ----
    "dist_near": 1000,   # ≤此值绿色通过（同一园区/街道）
    "dist_mid": 5000,    # 1km-5km黄色需人工

    # ---- 第2层：搜索降级域名白名单（LLM不可用时） ----
    "fallback_authoritative_domains": [
        "gov.cn",
        "baike.baidu.com",
    ],
}


# ============================================================
# 地址清洗
# ============================================================
def clean_address(addr: str) -> str:
    """
    清洗地址：去行政前缀、统一标点、去空白、统一常见缩写。

    目标：让"山东省济南市工业南路89号"和"济南市工业南路89号"
          在清洗后都变成"济南工业南路89号"，提升文本匹配准确率。
    """
    if not isinstance(addr, str) or not addr.strip():
        return ""

    result = addr.strip()

    # 1. 统一全角标点为半角
    result = result.replace("（", "(").replace("）", ")")
    result = result.replace("，", ",").replace("。", ".")
    result = result.replace("；", ";").replace("：", ":")
    result = result.replace("－", "-").replace("—", "-")
    result = result.replace("’", "'").replace("‘", "'")
    result = result.replace("“", '"').replace("”", '"')

    # 2. 去除多余空白
    result = re.sub(r'\s+', '', result)

    # 3. 去除省/自治区前缀
    result = re.sub(
        r'^(北京市|天津市|上海市|重庆市|'
        r'河北省|山西省|辽宁省|吉林省|黑龙江省|'
        r'江苏省|浙江省|安徽省|福建省|江西省|山东省|'
        r'河南省|湖北省|湖南省|广东省|海南省|'
        r'四川省|贵州省|云南省|陕西省|甘肃省|青海省|'
        r'台湾省|'
        r'内蒙古自治区|广西壮族自治区|西藏自治区|宁夏回族自治区|新疆维吾尔自治区|'
        r'香港特别行政区|澳门特别行政区)',
        '', result
    )

    # 4. 统一常见缩写
    abbreviations = {
        "高新技术产业开发区": "高新区",
        "经济技术开发区": "开发区",
        "化学工业区": "化工区",
        "保税港区": "保税区",
        "出口加工区": "出口加工区",  # 保持不变
    }
    for full, abbr in abbreviations.items():
        result = result.replace(full, abbr)

    # 5. 去除连续的"市/区/县/州/盟/旗"层级前缀
    #    但保留"街道/镇/乡/路/街/号/栋/座/层/楼"之后的门牌号内容
    #    策略：反复去掉开头的"XX市"、"XX区"、"XX县"等
    prev = None
    while prev != result:
        prev = result
        result = re.sub(
            r'^([一-龥]{2,6}(?:市|区|县|州|盟|旗|自治州|自治县|自治旗))',
            '', result
        )

    return result


# ============================================================
# 第1层：多策略文本相似度
# ============================================================
def multi_strategy_similarity(addr1: str, addr2: str) -> Tuple[float, str]:
    """
    用三种策略计算地址相似度，取最高分。

    策略：
      - ratio:          全字符串编辑距离（精确但敏感）
      - token_sort_ratio: 分词后排序再比较（容忍词序差异）
      - partial_ratio:   子串匹配（容忍地址冗余前缀）

    返回：(最高分数, 策略名称)
    """
    s1 = clean_address(addr1)
    s2 = clean_address(addr2)

    if not s1 or not s2:
        return 0.0, "ratio"

    strategies = [
        ("ratio", fuzz.ratio(s1, s2)),
        ("token_sort", fuzz.token_sort_ratio(s1, s2)),
        ("partial", fuzz.partial_ratio(s1, s2)),
    ]

    best_name, best_score = max(strategies, key=lambda x: x[1])
    return float(best_score), best_name


# ============================================================
# 第1.5层：地理编码距离（保留原逻辑）
# ============================================================
def geocode(address: str) -> Optional[Tuple[float, float]]:
    """调腾讯地图API将地址转成经纬度，返回 (lng, lat) 或 None"""
    if not CONFIG["tencent_map_key"]:
        return None

    url = "https://apis.map.qq.com/ws/geocoder/v1/"
    params = {"address": address, "key": CONFIG["tencent_map_key"]}

    try:
        resp = requests.get(url, params=params, timeout=5)
        data = resp.json()
        if data.get("status") == 0:
            loc = data["result"]["location"]
            return (loc["lng"], loc["lat"])
    except Exception:
        pass

    return None


def haversine_distance(coord1: Tuple[float, float],
                       coord2: Tuple[float, float]) -> float:
    """Haversine公式算两点间直线距离（米）"""
    lng1, lat1 = coord1
    lng2, lat2 = coord2
    R = 6371000  # 地球半径（米）

    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lng2 - lng1)

    a = (math.sin(delta_phi / 2) ** 2
         + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2)
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    return R * c


def check_geocode_distance(addr1: str, addr2: str):
    """
    调地图API算两个地址的直线距离。
    返回：(距离米数或None, 坐标1, 坐标2)
    """
    coord1 = geocode(addr1)
    coord2 = geocode(addr2)

    if coord1 is None or coord2 is None:
        return None, coord1, coord2

    distance = haversine_distance(coord1, coord2)
    return distance, coord1, coord2


# ============================================================
# 第2层：Web Search + LLM内容评估
# ============================================================
def _run_anysearch(query: str, max_results: int = 2) -> str:
    """调AnySearch CLI搜索，返回原始输出文本。"""
    import subprocess

    cli_path = CONFIG.get("anysearch_cli", "")
    if not cli_path:
        print("  ⚠️ AnySearch CLI 路径未配置（设置环境变量 ANYSEARCH_CLI）")
        return ""

    cmd = [
        sys.executable,
        cli_path,
        "search", query,
        "--max_results", str(max_results),
        "--zone", "cn",
    ]
    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=30
        )
        return result.stdout
    except subprocess.TimeoutExpired:
        print(f"  ⚠️ AnySearch 搜索超时 [{query[:40]}]")
        return ""
    except Exception as e:
        print(f"  ⚠️ AnySearch 调用失败: {e}")
        return ""


def _parse_search_results(raw_output: str) -> list[dict]:
    """从AnySearch输出中提取搜索结果（标题+URL+摘要）。"""
    results = []

    # 匹配URL（格式: - **URL**: https://...）
    urls = re.findall(r'\*\*URL\*\*: (https?://[^\s\n]+)', raw_output)
    # 匹配标题（格式: ### 1. 标题文本）
    titles = re.findall(r'### \d+\. (.+?)(?:\n|$)', raw_output)

    for i in range(len(urls)):
        result = {"url": urls[i]}
        if i < len(titles):
            result["title"] = titles[i].strip()
        results.append(result)

    return results


def _llm_evaluate(company: str, send_addr: str,
                  search_results: list[dict]) -> Optional[dict]:
    """
    调 DeepSeek API（OpenAI兼容接口）评估搜索结果能否佐证发函地址。

    返回：{"conclusion": "confirmed"|"suspicious"|"no_evidence",
           "reason": "...", "authoritative": bool, "address_mentioned": bool}
    如果API不可用或未配置Key，返回 None。
    """
    if not CONFIG.get("llm_api_key"):
        return None

    # 格式化搜索结果（只用标题+URL发给LLM）
    results_text = ""
    for i, r in enumerate(search_results[:5]):
        results_text += (
            f"[{i+1}] 标题: {r.get('title', '无标题')}\n"
            f"    URL: {r.get('url', '')}\n\n"
        )

    if not results_text.strip():
        return {
            "conclusion": "no_evidence",
            "reason": "搜索无结果",
            "authoritative": False,
            "address_mentioned": False,
        }

    prompt = f"""你是一个审计助理，评估以下搜索结果能否证明公司发函地址有效。

公司名：{company}
发函地址：{send_addr}

搜索结果：
{results_text}

请判断：
1. 这些结果是权威来源吗？（政府公告/公司官网/权威媒体/百度百科=权威；
   B2B平台/聚合信息站/自媒体=不权威）
2. 搜索结果是否明确提到了该公司的地址信息？
3. 综合判断：能佐证发函地址吗？

返回**纯JSON**（不要markdown代码块），格式：
{{"authoritative": true/false, "address_mentioned": true/false, "conclusion": "confirmed|suspicious|no_evidence", "reason": "一句话中文理由"}}"""

    try:
        resp = requests.post(
            f"{CONFIG['llm_base_url']}/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {CONFIG['llm_api_key']}",
                "Content-Type": "application/json",
            },
            json={
                "model": CONFIG["llm_model"],
                "messages": [
                    {"role": "system", "content": "你是一个审计助理。只返回JSON，不要加任何解释。"},
                    {"role": "user", "content": prompt},
                ],
                "max_tokens": 200,
                "temperature": 0.1,
            },
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        text = data["choices"][0]["message"]["content"].strip()

        # 清理可能的 markdown 代码块包裹
        if text.startswith("```"):
            text = re.sub(r'^```(?:json)?\s*', '', text)
            text = re.sub(r'\s*```$', '', text)

        result = json.loads(text)
        # 校验必要字段
        for key in ["conclusion", "reason", "authoritative", "address_mentioned"]:
            if key not in result:
                result[key] = False if key != "reason" else "LLM返回格式异常"
        return result

    except Exception as e:
        print(f"  ⚠️ LLM评估失败: {e}")
        return None


def _fallback_domain_check(search_results: list[dict]) -> dict:
    """
    降级策略：LLM不可用时，用收紧后的域名白名单判定。

    仅 gov.cn 和 baike.baidu.com 算权威来源。
    """
    domains = CONFIG["fallback_authoritative_domains"]
    has_auth = False
    matched_urls = []

    for r in search_results:
        url = r.get("url", "")
        for domain in domains:
            if domain in url:
                has_auth = True
                matched_urls.append(url)
                break

    if has_auth:
        return {
            "conclusion": "confirmed",
            "reason": f"域名白名单匹配（降级模式）: {', '.join(matched_urls[:2])}",
            "authoritative": True,
            "address_mentioned": True,
        }
    elif len(search_results) > 0:
        return {
            "conclusion": "suspicious",
            "reason": "降级模式：有搜索结果但非gov.cn/百度百科",
            "authoritative": False,
            "address_mentioned": False,
        }
    else:
        return {
            "conclusion": "no_evidence",
            "reason": "降级模式：搜索无结果",
            "authoritative": False,
            "address_mentioned": False,
        }


def search_and_evaluate(company: str, send_addr: str) -> dict:
    """
    第2层：搜索 + 评估。

    返回：{"conclusion": "confirmed"|"suspicious"|"no_evidence",
           "reason": "...",
           "authoritative": bool,
           "address_mentioned": bool,
           "mode": "llm"|"fallback",
           "search_snippet": str}
    """
    # 构建多条搜索策略
    queries = [
        f'"{company}" "{send_addr[:10]}"',
        f'"{company}" "{send_addr[:6]}" 项目',
        f'"{company}" 地址',
    ]

    all_results = []
    for query in queries:
        raw = _run_anysearch(query, max_results=2)
        parsed = _parse_search_results(raw)
        all_results.extend(parsed)

    # 去重（按URL）
    seen = set()
    unique_results = []
    for r in all_results:
        url = r.get("url", "")
        if url and url not in seen:
            seen.add(url)
            unique_results.append(r)

    # 生成搜索摘要
    snippet_parts = []
    for r in unique_results[:5]:
        snippet_parts.append(f"- {r.get('title', '?')}: {r.get('url', '?')}")
    search_snippet = "\n".join(snippet_parts) if snippet_parts else "无搜索结果"

    # 先尝试LLM评估
    llm_result = _llm_evaluate(company, send_addr, unique_results)

    if llm_result is not None:
        llm_result["mode"] = "llm"
        llm_result["search_snippet"] = search_snippet
        return llm_result

    # 降级为域名白名单
    fallback = _fallback_domain_check(unique_results)
    fallback["mode"] = "fallback"
    fallback["search_snippet"] = search_snippet
    return fallback


# ============================================================
# 风险评分
# ============================================================
def calculate_risk_score(sim_score: float,
                         distance: Optional[float],
                         search_result: Optional[dict]) -> Tuple[int, str]:
    """
    综合地址相似度、距离、搜索结果打分。

    返回：(0-100分, "低风险"|"中风险"|"高风险")
    """
    score = 0

    # --- 地址相似度信号 ---
    if sim_score < 60:
        score += 25
    elif sim_score < CONFIG["sim_mid"]:
        score += 15
    elif sim_score < CONFIG["sim_high"]:
        score += 5
    # sim ≥ 85: +0

    # --- 距离信号 ---
    if distance is not None:
        if distance > 10000:
            score += 20
        elif distance > CONFIG["dist_mid"]:
            score += 10
        # ≤5km: +0
    else:
        # 无法获取距离，轻微惩罚
        score += 5

    # --- 搜索信号 ---
    if search_result is not None:
        conclusion = search_result.get("conclusion", "")
        if conclusion == "no_evidence":
            score += 25      # 搜索无证据（不单独决定标红，需其他信号配合）
        elif conclusion == "suspicious":
            score += 15
        elif conclusion == "confirmed":
            score += 0
        else:
            score += 10
    else:
        # 搜索完全失败
        score += 15

    # --- 映射到风险等级 ---
    score = min(score, 100)
    if score <= 25:
        level = "低风险"
    elif score <= 50:
        level = "中风险"
    else:
        level = "高风险"

    return score, level


# ============================================================
# 综合判定
# ============================================================
def determine_verdict(sim_score: float,
                      distance: Optional[float],
                      dist_passed_near: bool,
                      dist_passed_mid: bool,
                      search_result: Optional[dict],
                      risk_level: str) -> Tuple[str, str]:
    """
    综合所有信号输出核查结论和判定理由。

    返回：(结论, 理由)
    结论："通过" | "需人工判断" | "异常"
    """
    reasons = []

    # 第1层判定
    if sim_score >= CONFIG["sim_high"]:
        reasons.append(f"地址高度相似({sim_score}%)")
        return "通过", " | ".join(reasons)

    # 第1.5层判定（进入此分支说明 sim < 85）
    if distance is not None:
        if dist_passed_near:
            reasons.append(f"距离极近({round(distance)}m)，地址相似度{sim_score}%")
            return "通过", " | ".join(reasons)
        elif dist_passed_mid:
            reasons.append(f"距离较近({round(distance)}m)但地址相似度偏低({sim_score}%)")
            # 即使距离中等，如果相似度很低也不能直接放过
            if sim_score >= CONFIG["sim_mid"]:
                return "需人工判断", " | ".join(reasons)
            # sim < 70，继续往下走

    # 距离失败/太远/相似度太低，进入第2层判定
    if search_result is None:
        reasons.append("搜索评估不可用")
        if risk_level == "低风险":
            return "需人工判断", " | ".join(reasons)
        else:
            return "异常", " | ".join(reasons)

    conclusion = search_result.get("conclusion", "")
    search_reason = search_result.get("reason", "")

    if conclusion == "confirmed":
        reasons.append(f"搜索佐证通过: {search_reason}")
        return "通过", " | ".join(reasons)
    elif conclusion == "suspicious":
        reasons.append(f"搜索来源可疑: {search_reason}")
        return "需人工判断", " | ".join(reasons)
    elif conclusion == "no_evidence":
        reasons.append(f"无权威来源佐证: {search_reason}")
        # 仅当其他信号也强烈异常时才直接标红；否则留给人审
        if risk_level == "高风险" or sim_score < 60:
            return "异常", " | ".join(reasons)
        else:
            return "需人工判断", " | ".join(reasons)

    # fallback
    reasons.append(f"搜索结论不明: {conclusion}")
    return "需人工判断", " | ".join(reasons)


# ============================================================
# 主流程
# ============================================================
def verify_addresses(input_file, output_file=None, skip_search=False):
    """
    三层过滤主函数（三色分级版）

    Args:
        input_file: 输入Excel路径
        output_file: 输出Excel路径（可选）
        skip_search: True时跳过第2层搜索，仅用第1层+第1.5层
    """
    # 读入数据
    if str(input_file).endswith('.csv'):
        df = pd.read_csv(input_file)
    else:
        df = pd.read_excel(input_file)

    # 检查必要的列（先精确匹配，再模糊匹配）
    required_cols = ["公司名称", "发函地址", "工商注册地址"]
    col_map = {}
    for c in required_cols:
        # 精确匹配优先
        if c in df.columns:
            col_map[c] = c
            continue
        # 模糊匹配：列名包含关键词
        for col in df.columns:
            if c in col:
                col_map[c] = col
                break

    if len(col_map) == len(required_cols):
        df = df.rename(columns={v: k for k, v in col_map.items()})
    else:
        missing = [c for c in required_cols if c not in col_map]
        print(f"❌ 缺少列: {missing}")
        print(f"   现有列: {list(df.columns)}")
        print("   请确保Excel包含：公司名称 | 发函地址 | 工商注册地址")
        return None

    # 处理空数据
    df["公司名称"] = df["公司名称"].fillna("").astype(str)
    df["发函地址"] = df["发函地址"].fillna("").astype(str)
    df["工商注册地址"] = df["工商注册地址"].fillna("").astype(str)

    print(f"📄 读取到 {len(df)} 条记录，开始核查...\n")

    # 初始化结果列
    df["清洗后_发函"] = ""
    df["清洗后_工商"] = ""
    df["相似度(%)"] = 0.0
    df["匹配策略"] = ""
    df["距离(米)"] = ""
    df["搜索来源"] = ""
    df["搜索评估"] = ""
    df["风险评分"] = 0
    df["风险等级"] = ""
    df["判定理由"] = ""
    df["核查结论"] = ""

    total = len(df)
    stats = {"通过": 0, "需人工判断": 0, "异常": 0}

    for idx, row in tqdm(df.iterrows(), total=total, desc="核查进度", ncols=80):
        company = str(row["公司名称"])
        send_addr = str(row["发函地址"])
        reg_addr = str(row["工商注册地址"])

        # 跳过空行
        if not company or company == "nan":
            continue

        print(f"  {company}")

        # ---- 清洗地址 ----
        clean_send = clean_address(send_addr)
        clean_reg = clean_address(reg_addr)
        df.at[idx, "清洗后_发函"] = clean_send
        df.at[idx, "清洗后_工商"] = clean_reg

        # ---- 第1层：多策略相似度 ----
        sim_score, sim_strategy = multi_strategy_similarity(send_addr, reg_addr)
        df.at[idx, "相似度(%)"] = sim_score
        df.at[idx, "匹配策略"] = sim_strategy

        # ---- 初始化第1.5层和第2层结果 ----
        distance = None
        dist_passed_near = False   # ≤1km
        dist_passed_mid = False    # ≤5km
        search_result = None

        # ---- 第1层判定：≥85% 直接通过 ----
        if sim_score >= CONFIG["sim_high"]:
            df.at[idx, "距离(米)"] = "0"
            df.at[idx, "搜索来源"] = "无需搜索"
            df.at[idx, "搜索评估"] = "N/A"
            df.at[idx, "风险评分"] = 0
            df.at[idx, "风险等级"] = "低风险"
            df.at[idx, "判定理由"] = f"地址高度相似({sim_score}%, {sim_strategy})"
            df.at[idx, "核查结论"] = "通过"
            stats["通过"] += 1
            print(f"  ✅ 第1层通过 (相似度: {sim_score}%, 策略: {sim_strategy})")
            continue

        # ---- 没通过第1层，进第1.5层 ----
        print(f"  ➡️ 第1层未通过 (相似度: {sim_score}%)")

        distance, coord1, coord2 = check_geocode_distance(send_addr, reg_addr)

        if distance is not None:
            df.at[idx, "距离(米)"] = round(distance)
            dist_passed_near = distance <= CONFIG["dist_near"]
            dist_passed_mid = distance <= CONFIG["dist_mid"]
        else:
            df.at[idx, "距离(米)"] = "API失败"

        # ---- 第1.5层判定 ----
        if dist_passed_near:
            df.at[idx, "搜索来源"] = "无需搜索"
            df.at[idx, "搜索评估"] = "N/A"
            score, level = calculate_risk_score(sim_score, distance, None)
            df.at[idx, "风险评分"] = score
            df.at[idx, "风险等级"] = level
            verdict, reason = determine_verdict(
                sim_score, distance,
                dist_passed_near=True, dist_passed_mid=True,
                search_result=None, risk_level=level
            )
            df.at[idx, "判定理由"] = reason
            df.at[idx, "核查结论"] = verdict
            stats[verdict] = stats.get(verdict, 0) + 1
            print(f"  ✅ 第1.5层通过 (距离: {round(distance)}m)")
            continue

        if dist_passed_mid and sim_score >= CONFIG["sim_mid"]:
            # 距离中等但相似度还过得去 → 黄色
            df.at[idx, "搜索来源"] = "跳过搜索"
            df.at[idx, "搜索评估"] = "N/A"
            score, level = calculate_risk_score(sim_score, distance, None)
            df.at[idx, "风险评分"] = score
            df.at[idx, "风险等级"] = level
            df.at[idx, "判定理由"] = (
                f"距离较近({round(distance)}m)但地址相似度偏低({sim_score}%)"
            )
            df.at[idx, "核查结论"] = "需人工判断"
            stats["需人工判断"] += 1
            print(f"  ⚠️ 需人工判断 (距离: {round(distance)}m, 相似度: {sim_score}%)")
            continue

        # ---- 进入第2层：搜索 + 评估 ----
        if skip_search:
            score, level = calculate_risk_score(sim_score, distance, None)
            df.at[idx, "搜索来源"] = "已跳过"
            df.at[idx, "搜索评估"] = "N/A"
            df.at[idx, "风险评分"] = score
            df.at[idx, "风险等级"] = level
            # 无搜索佐证时，根据地址+距离信号判定
            if sim_score >= CONFIG["sim_mid"]:
                df.at[idx, "核查结论"] = "需人工判断"
                df.at[idx, "判定理由"] = "跳过搜索，地址相似度中等，建议人工复核"
            else:
                df.at[idx, "核查结论"] = "异常"
                df.at[idx, "判定理由"] = "跳过搜索，地址相似度低且距离远"
            verdict = df.at[idx, "核查结论"]
            stats[verdict] = stats.get(verdict, 0) + 1
            icon = {"通过": "✅", "需人工判断": "⚠️", "异常": "❌"}
            print(f"  {icon.get(verdict, '?')} {verdict} (搜索已跳过)")
            print()
            continue

        dist_str = f"{round(distance)}m" if distance else "N/A"
        print(f"  🔍 距离: {dist_str}，进入第2层搜索评估...")

        search_result = search_and_evaluate(company, send_addr)
        df.at[idx, "搜索来源"] = search_result.get("mode", "fallback")
        df.at[idx, "搜索评估"] = search_result.get("reason", "")

        score, level = calculate_risk_score(sim_score, distance, search_result)
        df.at[idx, "风险评分"] = score
        df.at[idx, "风险等级"] = level

        verdict, reason = determine_verdict(
            sim_score, distance,
            dist_passed_near=False,
            dist_passed_mid=dist_passed_mid,
            search_result=search_result,
            risk_level=level
        )
        df.at[idx, "判定理由"] = reason
        df.at[idx, "核查结论"] = verdict
        stats[verdict] = stats.get(verdict, 0) + 1

        icon = {"通过": "✅", "需人工判断": "⚠️", "异常": "❌"}
        print(f"  {icon.get(verdict, '?')} {verdict}: {reason}")

        print()

    # ---- 输出报告 ----
    if output_file is None:
        input_path = Path(input_file)
        output_file = input_path.parent / f"{input_path.stem}_核查结果{input_path.suffix}"

    # 定义输出列顺序
    output_columns = [
        "公司名称", "发函地址", "工商注册地址",
        "清洗后_发函", "清洗后_工商",
        "相似度(%)", "匹配策略",
        "距离(米)",
        "搜索来源", "搜索评估",
        "风险评分", "风险等级",
        "判定理由", "核查结论",
    ]
    # 只保留存在的列
    output_columns = [c for c in output_columns if c in df.columns]
    df_out = df[output_columns]

    # 写入Excel
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    df_out.to_excel(writer, index=False, sheet_name='核查结果')
    ws = writer.sheets['核查结果']

    # 列宽
    col_widths = {
        "公司名称": 22, "发函地址": 32, "工商注册地址": 32,
        "清洗后_发函": 28, "清洗后_工商": 28,
        "相似度(%)": 10, "匹配策略": 12,
        "距离(米)": 10,
        "搜索来源": 10, "搜索评估": 40,
        "风险评分": 8, "风险等级": 8,
        "判定理由": 45, "核查结论": 12,
    }
    for col_idx, col_name in enumerate(output_columns, 1):
        width = col_widths.get(col_name, 12)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 条件格式：核查结论列三色标记
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_font = Font(color="006100")
    yellow_font = Font(color="9C6500")
    red_font = Font(color="9C0006")

    conclusion_col = len(output_columns)
    col_letter = get_column_letter(conclusion_col)

    for row_idx in range(2, len(df_out) + 2):
        cell = ws[f"{col_letter}{row_idx}"]
        if cell.value == "通过":
            cell.fill = green_fill
            cell.font = green_font
        elif cell.value == "需人工判断":
            cell.fill = yellow_fill
            cell.font = yellow_font
        elif cell.value == "异常":
            cell.fill = red_fill
            cell.font = red_font

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选器
    ws.auto_filter.ref = f"A1:{get_column_letter(len(output_columns))}{len(df_out)+1}"

    # 风险等级列也加颜色
    risk_col_idx = output_columns.index("风险等级") + 1
    risk_letter = get_column_letter(risk_col_idx)
    for row_idx in range(2, len(df_out) + 2):
        cell = ws[f"{risk_letter}{row_idx}"]
        if cell.value == "低风险":
            cell.fill = green_fill
            cell.font = green_font
        elif cell.value == "中风险":
            cell.fill = yellow_fill
            cell.font = yellow_font
        elif cell.value == "高风险":
            cell.fill = red_fill
            cell.font = red_font

    writer.close()

    # 统计输出
    print(f"\n{'='*55}")
    print(f"📊 核查完成")
    print(f"   总数: {total}")
    print(f"   ✅ 通过: {stats.get('通过', 0)} 条 (可直接发函)")
    print(f"   ⚠️  需人工判断: {stats.get('需人工判断', 0)} 条 (建议实习生优先查)")
    print(f"   ❌ 异常: {stats.get('异常', 0)} 条 (重点风险)")
    print(f"   报告已保存: {output_file}")

    return df


# ============================================================
# 入口
# ============================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="函证地址核查 - 三层过滤异常初筛器"
    )
    parser.add_argument(
        "input",
        help="输入Excel路径（含列：公司名称 | 发函地址 | 工商注册地址）"
    )
    parser.add_argument(
        "-o", "--output",
        help="输出Excel路径（默认：输入_核查结果.xlsx）"
    )
    parser.add_argument(
        "--map-key",
        help="腾讯地图API Key"
    )
    parser.add_argument(
        "--llm-key",
        help="LLM API Key（DeepSeek/OpenAI兼容接口，不提供则使用域名白名单降级）"
    )
    parser.add_argument(
        "--llm-model",
        default=None,
        help="LLM模型名称（默认: deepseek-chat）"
    )
    parser.add_argument(
        "--no-search",
        action="store_true",
        help="跳过第2层搜索，仅使用第1层+第1.5层"
    )
    args = parser.parse_args()

    if args.map_key:
        CONFIG["tencent_map_key"] = args.map_key
    if args.llm_key:
        CONFIG["llm_api_key"] = args.llm_key
    if args.llm_model:
        CONFIG["llm_model"] = args.llm_model

    if not CONFIG["tencent_map_key"]:
        print("⚠️ 未配置腾讯地图API Key")
        print("   第1.5层（地理编码）将跳过，仅使用第1层 + 第2层\n")

    if not CONFIG["llm_api_key"]:
        print("ℹ️ 未配置LLM API Key")
        print("   第2层将使用域名白名单降级模式（仅 gov.cn + 百度百科）\n")

    df = verify_addresses(args.input, args.output, skip_search=args.no_search)
