"""固定资产折旧年审测算引擎。

支持交互和非交互两种模式。
"""

from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from audit_tools.common.amount import to_number, parse_date
from audit_tools.common.logging import get_logger
from audit_tools.depreciation.helpers import normalize_life, expected_months
from audit_tools.depreciation.config import build_config, build_config_from_args

logger = get_logger(__name__)


def _cell(row, config, key, default=None):
    col = config.get(key)
    if col is None or col >= len(row):
        return default
    return row[col]


def row_to_asset(row, config):
    """将一行数据转为资产字典。"""
    orig_value = to_number(_cell(row, config, "orig_value"))
    residual = to_number(_cell(row, config, "residual"))
    residual_rate = to_number(_cell(row, config, "residual_rate"))
    if residual == 0 and residual_rate:
        residual = round(orig_value * residual_rate / 100, 2)

    return {
        "name": str(_cell(row, config, "name", "") or "").strip(),
        "entry_date": _cell(row, config, "entry_date"),
        "life": normalize_life(to_number(_cell(row, config, "life"))),
        "used_life_min": normalize_life(to_number(_cell(row, config, "used_life"))),
        "orig_value": orig_value,
        "residual": residual,
        "residual_rate": residual_rate,
        "monthly_rate": to_number(_cell(row, config, "monthly_rate")),
        "total_monthly": to_number(_cell(row, config, "monthly_depr")),
        "rows": 1,
        "disposed": to_number(_cell(row, config, "orig_decrease")) > 0,
        "disposal_month": 0,
    }


def collect_items(ws, config):
    """从工作表收集资产数据。"""
    rows = list(ws.iter_rows(min_row=config["data_start"], values_only=True))
    if not config.get("grouped"):
        items = []
        for index, row in enumerate(rows, 1):
            card = str(_cell(row, config, "card", index) or index).strip()
            name = str(_cell(row, config, "name", "") or "").strip()
            if not card and not name:
                continue
            items.append((card, row_to_asset(row, config)))
        return items

    groups = defaultdict(lambda: {
        "name": "", "entry_date": None, "life": 0, "used_life_min": None,
        "orig_value": 0, "residual": 0, "residual_rate": 0,
        "monthly_rate": 0, "total_monthly": 0, "rows": 0,
        "disposed": False, "disposal_month": 0,
    })

    for index, row in enumerate(rows, 1):
        card = str(_cell(row, config, "card", index) or index).strip()
        name = str(_cell(row, config, "name", "") or "").strip()
        if not card and not name:
            continue
        asset = row_to_asset(row, config)
        group = groups[card]
        group["name"] = asset["name"] or group["name"]
        group["entry_date"] = asset["entry_date"] or group["entry_date"]
        group["life"] = asset["life"] or group["life"]
        group["orig_value"] = asset["orig_value"] or group["orig_value"]
        group["residual"] = asset["residual"] or group["residual"]
        group["residual_rate"] = asset["residual_rate"] or group["residual_rate"]
        group["monthly_rate"] = asset["monthly_rate"] or group["monthly_rate"]
        group["total_monthly"] += asset["total_monthly"]
        group["rows"] += 1

        used = asset["used_life_min"]
        if used:
            if group["used_life_min"] is None or used < group["used_life_min"]:
                group["used_life_min"] = used

        if asset["disposed"]:
            group["disposed"] = True
            group["disposal_month"] = group["rows"]

    return sorted(groups.items(), key=lambda item: item[0])


def calculate(items, config):
    """执行折旧测算。"""
    audit_year = int(config["audit_year"])
    results = []
    skipped = {"zero_value": 0, "no_date": 0, "invalid_life": 0}
    match = 0
    mismatch = 0

    for card, asset in items:
        orig_value = asset["orig_value"]
        life = asset["life"]
        if orig_value == 0:
            skipped["zero_value"] += 1
            continue
        if life <= 0:
            skipped["invalid_life"] += 1
            continue

        acq_date = parse_date(asset["entry_date"])
        if acq_date is None:
            skipped["no_date"] += 1
            continue

        if config.get("method") == "rate" and asset.get("monthly_rate"):
            monthly_depr = round(orig_value * asset["monthly_rate"] / 100, 2)
        else:
            monthly_depr = round((orig_value - asset["residual"]) / life, 2)

        months, tag = expected_months(
            acq_date, audit_year, life, asset["used_life_min"],
            asset["disposed"], asset["disposal_month"], config.get("has_disposal"),
        )
        if tag == "未来年度入账":
            continue

        calc_total = round(monthly_depr * months, 2)
        customer_total = round(asset["total_monthly"], 2)
        diff = round(calc_total - customer_total, 2)
        result_status = "一致" if abs(diff) < 0.01 else "有差异"
        status = f"{tag},{result_status}" if tag else result_status

        if result_status == "一致":
            match += 1
        else:
            mismatch += 1

        customer_months = asset["rows"]
        month_status = "一致" if customer_months == months else f"应提{months}月，客户{customer_months}月"

        results.append([
            card, asset["name"], acq_date.strftime("%Y-%m-%d"),
            orig_value, asset["residual"], int(life),
            int(asset["used_life_min"]) if asset["used_life_min"] else "",
            customer_months, months, monthly_depr,
            calc_total, customer_total, diff, status, month_status,
        ])

    return results, match, mismatch, skipped


def write_results(path, results):
    """输出 Excel 结果文件。"""
    out_path = path.with_name(f"{path.stem}_折旧年审结果.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "折旧测算"

    headers = [
        "卡片编码", "资产名称", "入账日期", "期初原值", "残值",
        "使用寿命(月)", "年初已提月数", "客户实际月数", "应提月数",
        "月折旧额", "测算年折旧额", "客户年折旧额(汇总)", "差异",
        "结论", "月数核对",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results) + 1}"

    red_font = Font(color="FF0000")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell_obj in ws[1]:
        cell_obj.font = Font(bold=True)
        cell_obj.fill = header_fill

    for row in results:
        ws.append(row)
        if "有差异" in str(row[13]):
            for cell_obj in ws[ws.max_row]:
                cell_obj.font = red_font

    diff_ws = wb.create_sheet(title="差异汇总")
    diff_ws.append(["卡片编码", "资产名称", "测算年折旧额", "客户年折旧额", "差异", "备注"])
    total_diff = 0
    for row in results:
        if "有差异" not in str(row[13]):
            continue
        diff_ws.append([row[0], row[1], row[10], row[11], row[12], row[13]])
        total_diff += abs(row[12] or 0)

    diff_ws.append([])
    diff_ws.append(["差异绝对值合计", "", "", "", round(total_diff, 2), ""])
    diff_ws.freeze_panes = "A2"

    for sheet in [ws, diff_ws]:
        for col_idx in range(1, sheet.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = max(
                len(str(sheet.cell(row=row_idx, column=col_idx).value or ""))
                for row_idx in range(1, sheet.max_row + 1)
            )
            sheet.column_dimensions[letter].width = min(max(max_len + 2, 10), 30)

    wb.save(out_path)
    return out_path


def run_interactive():
    """交互式运行（原始工作流，向后兼容）。"""
    print("请把客户固定资产明细 Excel 拖进终端，然后回车：")
    path = Path(input().strip().strip("'\""))

    if not path.exists():
        raise SystemExit(f"找不到文件：{path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    config_dir = Path(__file__).resolve().parent.parent.parent.parent / "客户配置"
    customer_name = path.stem
    config_path = config_dir / f"{customer_name}.json"
    config = build_config(ws, config_path, customer_name)

    _run(path, config, ws)


def run_non_interactive(
    input_file,
    output_file=None,
    config_file=None,
    audit_year=2025,
    method="straight",
    grouped=False,
    has_disposal=False,
    data_start=None,
    header_row=None,
):
    """非交互式运行（CLI 模式）。"""
    path = Path(input_file)
    if not path.exists():
        raise FileNotFoundError(f"找不到文件：{path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    config_path = Path(config_file) if config_file else None
    config = build_config_from_args(
        ws,
        audit_year=audit_year,
        method=method,
        grouped=grouped,
        has_disposal=has_disposal,
        data_start=data_start,
        header_row_idx=header_row,
        config_path=config_path,
    )

    return _run(path, config, ws, output_file)


def _run(path, config, ws, output_file=None):
    """执行测算并输出。"""
    logger.info("正在测算 %d 年折旧...", config["audit_year"])
    items = collect_items(ws, config)
    results, match, mismatch, skipped = calculate(items, config)

    out_path = write_results(path, results)

    logger.info("完成：%s", out_path)
    logger.info("共处理 %d 条资产。一致：%d；有差异：%d。", len(results), match, mismatch)
    logger.info(
        "跳过：原值为 0 %d 条；无有效日期 %d 条；寿命异常 %d 条。",
        skipped["zero_value"], skipped["no_date"], skipped["invalid_life"],
    )

    return out_path
