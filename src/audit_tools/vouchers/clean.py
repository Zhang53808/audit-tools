"""凭证明细 Excel 清洗。

功能:
  - 自动查找表头行
  - 跳过空行、合计行
  - 统一月份格式
  - 清洗金额字段
"""

import re
from pathlib import Path

import openpyxl

from audit_tools.common.amount import clean_amount
from audit_tools.common.logging import get_logger

logger = get_logger(__name__)


def clean_month(value):
    """统一月份格式：2025年3月、03月、3 → 3月。"""
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""

    patterns = [
        r"年\s*(\d{1,2})\s*月",
        r"(\d{1,2})\s*月",
        r"^(\d{1,2})$",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            month = int(match.group(1))
            if 1 <= month <= 12:
                return f"{month}月"
    return text


def find_header_row(rows):
    """在 rows 中定位表头行。"""
    for index, row in enumerate(rows):
        text = " ".join(str(value) for value in row if value is not None)
        if "序号" in text and "月份" in text and "凭证" in text:
            return index
        if "月" in text and "凭证" in text and "金额" in text:
            return index
    return None


def build_column_index(headers):
    """根据表头构建列索引。"""
    columns = {}
    for col, header in headers:
        if "月" in header:
            columns["month"] = col
        elif "凭证字" in header:
            columns["type"] = col
        elif "凭证号" in header:
            columns["num"] = col
        elif "金额" in header:
            columns["amount"] = col
        elif "备注" in header or "摘要" in header:
            columns["note"] = col
    return columns


def process(input_path: str, output_path: str = None) -> str:
    """清洗凭证明细并输出。

    Args:
        input_path: 输入 Excel 路径
        output_path: 输出路径（默认：输入_已清洗.xlsx）

    Returns:
        输出文件路径
    """
    path = Path(input_path)
    if not path.exists():
        raise FileNotFoundError(f"找不到文件：{path}")

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]

    header_idx = find_header_row(rows)
    if header_idx is None:
        raise SystemExit("找不到表头行，请确认表格里包含月份、凭证、金额等字段。")

    headers = []
    for col, value in enumerate(rows[header_idx]):
        if value is None:
            continue
        text = str(value).strip()
        if text:
            headers.append((col, text))

    col_index = build_column_index(headers)
    logger.info("识别到的列：%s", col_index)

    cleaned = []
    for row in rows[header_idx + 1:]:
        if all(value is None or str(value).strip() == "" for value in row):
            continue
        if any("合计" in str(value) for value in row if value is not None):
            continue
        if row[0] is None or str(row[0]).strip() == "":
            continue

        new_row = list(row)
        if "month" in col_index:
            new_row[col_index["month"]] = clean_month(new_row[col_index["month"]])
        if "amount" in col_index:
            new_row[col_index["amount"]] = clean_amount(new_row[col_index["amount"]])
        cleaned.append(new_row)

    if output_path is None:
        output_path = str(path.with_name(f"{path.stem}_已清洗.xlsx"))

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "清洗结果"

    for col, header in headers:
        out_ws.cell(row=1, column=col + 1, value=header)

    for row_idx, row in enumerate(cleaned, 2):
        for col_idx, value in enumerate(row, 1):
            out_ws.cell(row=row_idx, column=col_idx, value=value)

    out_ws.freeze_panes = "A2"
    out_wb.save(output_path)

    logger.info("清洗完成：%s", output_path)
    logger.info("有效数据：%d 行", len(cleaned))
    for row in cleaned[:5]:
        month = row[col_index["month"]] if "month" in col_index else "?"
        amount = row[col_index["amount"]] if "amount" in col_index else "?"
        logger.info("  月份：%s    金额：%s", month, amount)

    return output_path
