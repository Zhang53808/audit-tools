"""Excel 输出格式化工具。

消除 address_verification.py 和 related_party_check.py 中 ~80行重复的 openpyxl 三色标记代码。
"""

from typing import Dict, List, Optional, Tuple, Union

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# 预置样式常量
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FONT = Font(color="006100")
YELLOW_FONT = Font(color="9C6500")
RED_FONT = Font(color="9C0006")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
RED_FONT_STYLE = Font(color="FF0000")

# 便捷映射
RISK_STYLE_MAP = {
    "低风险": (GREEN_FILL, GREEN_FONT),
    "中风险": (YELLOW_FILL, YELLOW_FONT),
    "高风险": (RED_FILL, RED_FONT),
}

CONCLUSION_STYLE_MAP = {
    "通过": (GREEN_FILL, GREEN_FONT),
    "需人工判断": (YELLOW_FILL, YELLOW_FONT),
    "异常": (RED_FILL, RED_FONT),
}


def apply_tricolor_column(
    ws: Worksheet,
    col_letter: str,
    data_rows: int,
    style_map: Dict[str, Tuple[PatternFill, Font]],
) -> None:
    """对指定列的单元格按值应用三色标记。

    Args:
        ws: openpyxl worksheet
        col_letter: 列字母 (如 "N")
        data_rows: 数据行数（不含表头）
        style_map: 值 -> (Fill, Font) 映射
    """
    for row_idx in range(2, data_rows + 2):
        cell = ws[f"{col_letter}{row_idx}"]
        if not cell.value:
            continue
        for label, (fill, font) in style_map.items():
            if label in str(cell.value):
                cell.fill = fill
                cell.font = font
                break


def apply_triggered_column(
    ws: Worksheet,
    col_letter: str,
    data_rows: int,
) -> None:
    """对触发标记列（非 ✗、非 —）的单元格标绿。"""
    for row_idx in range(2, data_rows + 2):
        cell = ws[f"{col_letter}{row_idx}"]
        if cell.value and str(cell.value) not in ("✗", "—"):
            cell.fill = GREEN_FILL
            cell.font = GREEN_FONT


def set_column_widths(ws: Worksheet, columns: List[str], width_map: Dict[str, int]) -> None:
    """按列名设置列宽。"""
    for col_idx, col_name in enumerate(columns, 1):
        if col_name in width_map:
            ws.column_dimensions[get_column_letter(col_idx)].width = width_map[col_name]


def style_header_row(ws: Worksheet, col_count: int) -> None:
    """表头行加粗 + 蓝底。"""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL


def write_dataframe_to_excel(
    df: pd.DataFrame,
    output_path: Union[str, any],
    sheet_name: str = "核查结果",
    col_widths: Optional[Dict[str, int]] = None,
    conclusion_col: Optional[str] = None,
    risk_col: Optional[str] = None,
    triggered_cols: Optional[List[str]] = None,
    freeze_pane: str = "A2",
    auto_filter: bool = True,
) -> Worksheet:
    """DataFrame -> 格式化 Excel 的一站式函数。

    Args:
        df: 要输出的 DataFrame
        output_path: 输出路径或 ExcelWriter
        sheet_name: 工作表名称
        col_widths: 列名 -> 宽度
        conclusion_col: 核查结论列名（自动三色标记）
        risk_col: 风险等级列名（自动三色标记）
        triggered_cols: 触发维度列名列表（非 ✗ 标绿）
        freeze_pane: 冻结窗格位置
        auto_filter: 是否启用自动筛选
    """
    writer = pd.ExcelWriter(output_path, engine='openpyxl') if isinstance(output_path, str) else output_path
    df.to_excel(writer, index=False, sheet_name=sheet_name)
    ws = writer.sheets[sheet_name]

    columns = list(df.columns)

    # 列宽
    if col_widths:
        set_column_widths(ws, columns, col_widths)

    # 冻结 + 筛选
    ws.freeze_panes = freeze_pane
    if auto_filter:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(df) + 1}"

    # 表头
    style_header_row(ws, len(columns))

    data_rows = len(df)

    # 核查结论列三色
    if conclusion_col and conclusion_col in columns:
        col_idx = columns.index(conclusion_col) + 1
        apply_tricolor_column(ws, get_column_letter(col_idx), data_rows, CONCLUSION_STYLE_MAP)

    # 风险等级列三色
    if risk_col and risk_col in columns:
        col_idx = columns.index(risk_col) + 1
        apply_tricolor_column(ws, get_column_letter(col_idx), data_rows, RISK_STYLE_MAP)

    # 触发维度列标绿
    if triggered_cols:
        for col_name in triggered_cols:
            if col_name in columns:
                col_idx = columns.index(col_name) + 1
                apply_triggered_column(ws, get_column_letter(col_idx), data_rows)

    if isinstance(output_path, str):
        writer.close()

    return ws
