#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
固定资产折旧测算脚本 —— 信沃医美专版（v1，已跑通）
列位置固定，无需配置，拖表即算。
"""

import openpyxl
from datetime import datetime
from collections import defaultdict

print("请把客户的固定资产明细表拖进终端，然后回车：")
path = input().strip().strip("'\"")

wb = openpyxl.load_workbook(path)
ws = wb.active

AUDIT_YEAR = 2025

COL_CARD = 8
COL_NAME = 9
COL_ENTRY = 12
COL_LIFE = 13
COL_USED = 14
COL_RESIDUAL = 16
COL_ORIG = 17
COL_MONTHLY = 28

rows_data = list(ws.iter_rows(min_row=2, values_only=True))

asset_groups = defaultdict(lambda: {
    "name": "",
    "entry_date": None,
    "life": 0,
    "used_life_min": None,
    "residual": 0,
    "orig_value": 0,
    "total_monthly": 0,
    "rows": 0
})

for row in rows_data:
    card = str(row[COL_CARD] or "").strip()
    name = str(row[COL_NAME] or "").strip()
    if not card and not name:
        continue
    g = asset_groups[card]
    g["name"] = name
    g["entry_date"] = row[COL_ENTRY]
    try:
        g["life"] = float(row[COL_LIFE]) if row[COL_LIFE] else 0
        used_val = float(row[COL_USED]) if row[COL_USED] else 0
        g["residual"] = float(row[COL_RESIDUAL]) if row[COL_RESIDUAL] else 0
        g["orig_value"] = float(row[COL_ORIG]) if row[COL_ORIG] else 0
        monthly = float(row[COL_MONTHLY]) if row[COL_MONTHLY] else 0
    except:
        continue
    if g["used_life_min"] is None or used_val < g["used_life_min"]:
        g["used_life_min"] = used_val
    g["total_monthly"] += monthly
    g["rows"] += 1

results = []
match = 0
mismatch = 0
zero_value = 0
no_date = 0

print(f"\n正在测算 {AUDIT_YEAR} 年折旧...\n")

for card, g in sorted(asset_groups.items()):
    name = g["name"]
    life = g["life"]
    used_life_min = g["used_life_min"]
    residual = g["residual"]
    orig_value = g["orig_value"]
    cust_total = round(g["total_monthly"], 2)

    if orig_value == 0:
        zero_value += 1
        continue
    if life <= 0:
        continue

    entry_date = g["entry_date"]
    acq_year, acq_month = AUDIT_YEAR, 1

    if isinstance(entry_date, datetime):
        acq_year, acq_month = entry_date.year, entry_date.month
    elif isinstance(entry_date, str):
        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%Y-%m", "%Y/%m", "%Y年%m月"]:
            try:
                dt = datetime.strptime(entry_date.strip(), fmt)
                acq_year, acq_month = dt.year, dt.month
                break
            except:
                pass
    else:
        no_date += 1
        continue

    monthly_depr = round((orig_value - residual) / life, 2)
    remaining_at_start = life - (used_life_min - 1) if used_life_min and used_life_min > 0 else life

    if remaining_at_start <= 0:
        calc_total, months = 0, 0
    else:
        if acq_year < AUDIT_YEAR:
            months = 12
        elif acq_year == AUDIT_YEAR:
            months = 12 - acq_month
            if months < 0:
                months = 0
        else:
            continue
        if months > remaining_at_start:
            months = int(remaining_at_start)
        if months <= 0:
            continue
        calc_total = round(monthly_depr * months, 2)

    diff = round(calc_total - cust_total, 2)
    status = "一致" if abs(diff) < 0.01 else "有差异"

    if status == "一致":
        match += 1
    else:
        mismatch += 1

    cust_months = g["rows"]
    month_status = "一致" if cust_months == months else f"应提{months}月,客户{cust_months}月"
    entry_str = str(entry_date) if entry_date else ""

    results.append([card, name, entry_str, orig_value, residual, int(life),
                   int(used_life_min) if used_life_min else "", cust_months, months,
                   monthly_depr, calc_total, cust_total, diff, status, month_status])

out_path = path.replace(".xlsx", "_折旧年审结果.xlsx")
out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = "折旧测算"

headers = ["卡片编码", "资产名称", "入账日期", "期初原值", "残值", "使用寿命(月)",
           "年初已提月数", "客户实际月数", "应提月数", "月折旧额",
           "测算年折旧额", "客户年折旧额(汇总)", "差异", "结论", "月数核对"]

for j, h in enumerate(headers):
    out_ws.cell(row=1, column=j+1, value=h)

red_font = openpyxl.styles.Font(color="FF0000")
for i, row in enumerate(results, 2):
    for j, v in enumerate(row):
        cell = out_ws.cell(row=i, column=j+1, value=v)
        if "差异" in str(row[-1]):
            cell.font = red_font

out_wb.save(out_path)

print(f"完成！结果：{out_path}")
print(f"共处理 {len(results)} 条资产")
print(f"  一致：{match}  有差异：{mismatch}  原值0：{zero_value}  无日期：{no_date}")
